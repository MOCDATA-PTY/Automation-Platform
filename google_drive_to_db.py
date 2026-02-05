import os
import io
import sys
import tempfile
from datetime import datetime
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

sys.stdout.reconfigure(line_buffering=True)

# Google Drive settings
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
CLIENT_SECRET_FILE = '/home/ethan/Desktop/Gork Test/Turn over GABE tuesday update/client_secret_929057555993-c86mkjhf08suobk6olcca6sgmudeg8l0.apps.googleusercontent.com.json'
TOKEN_FILE = '/home/ethan/Desktop/Gork Test/token.json'
FOLDER_ID = '1xPGTc8320el4HXmZ3tNsHcavGtZ4asIY'

# PostgreSQL settings
DB_HOST = "167.88.43.168"
DB_PORT = "5432"
DB_NAME = "turnover_data"
DB_USER = "powerbi"
DB_PASSWORD = "your_secure_password"

def get_drive_service():
    """Authenticate and return Google Drive service"""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=8080)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)

def list_folders(service, parent_id):
    """List folders in a parent folder"""
    query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def list_xlsx_files(service, folder_id):
    """List xlsx files in a folder"""
    query = f"'{folder_id}' in parents and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or name contains '.xlsx') and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def download_file(service, file_id):
    """Download a file and return as bytes"""
    request = service.files().get_media(fileId=file_id)
    file_buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    file_buffer.seek(0)
    return file_buffer

def get_branch(filename):
    name = filename.upper().replace('.XLSX', '')
    if 'ALL OVER' in name:
        return None
    return name

def parse_month_code(code):
    try:
        code_str = str(int(code))
        year = int(code_str[:4])
        month = int(code_str[4:])
        if 1 <= month <= 12:
            return datetime(year, month, 1)
    except:
        pass
    return None

def process_wide_format(sheet, branch):
    data = []
    month_columns = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row and len(row) > 6 and row[3] == 'Debtor':
            for col_idx in range(6, len(row)):
                if row[col_idx]:
                    date_val = parse_month_code(row[col_idx])
                    if date_val:
                        month_columns[col_idx] = date_val
            break

    if not month_columns:
        return data

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue
        debtor_code = row[3]
        debtor_name = row[4]
        if not debtor_code or debtor_code == 'Debtor':
            continue
        if not isinstance(debtor_code, str):
            continue
        for col_idx, date_val in month_columns.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    value = float(row[col_idx])
                    if value != 0:
                        data.append((
                            debtor_code,
                            debtor_name,
                            value,
                            date_val.strftime('%Y-%m-%d'),
                            date_val.strftime('%Y-%m-%d'),
                            branch
                        ))
                except (ValueError, TypeError):
                    continue
    return data

print("=" * 50)
print("GOOGLE DRIVE TO POSTGRESQL")
print("=" * 50)

# Authenticate with Google Drive
print("\n[GOOGLE] Authenticating...")
service = get_drive_service()
print("[GOOGLE] Authenticated!")

# List year folders
print(f"\n[GOOGLE] Scanning folder...")
year_folders = list_folders(service, FOLDER_ID)
print(f"[GOOGLE] Found {len(year_folders)} year folders: {[f['name'] for f in year_folders]}")

all_data = []

# Process each year folder
for year_folder in sorted(year_folders, key=lambda x: x['name']):
    year_name = year_folder['name']
    year_id = year_folder['id']

    xlsx_files = list_xlsx_files(service, year_id)
    xlsx_files = [f for f in xlsx_files if 'ALL OVER' not in f['name'].upper()]

    print(f"\n[YEAR {year_name}] Processing {len(xlsx_files)} files...")

    for file_info in xlsx_files:
        filename = file_info['name']
        file_id = file_info['id']
        branch = get_branch(filename)

        if branch is None:
            continue

        try:
            # Download file
            file_buffer = download_file(service, file_id)

            # Load workbook from buffer
            wb = load_workbook(file_buffer, data_only=True)
            sheet = wb.active

            # Process data
            data = process_wide_format(sheet, branch)
            all_data.extend(data)

            print(f"  -> {filename}: {len(data)} rows")
            wb.close()
        except Exception as e:
            print(f"  [ERROR] {filename}: {e}")

print(f"\n[TOTAL] {len(all_data)} rows collected")

# Push to PostgreSQL
print(f"\n[POSTGRESQL] Connecting to {DB_HOST}...")
conn = None
try:
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        connect_timeout=10
    )
    conn.autocommit = False
    cursor = conn.cursor()
    print("[POSTGRESQL] Connected!")

    print("[POSTGRESQL] Dropping old table...")
    cursor.execute("DROP TABLE IF EXISTS turnover_data CASCADE")
    conn.commit()

    print("[POSTGRESQL] Creating new table...")
    cursor.execute('''
        CREATE TABLE turnover_data (
            id SERIAL PRIMARY KEY,
            debtor VARCHAR(100),
            debtor_name VARCHAR(255),
            value DECIMAL(15,2),
            date DATE,
            date_fixed VARCHAR(20),
            branch VARCHAR(50)
        )
    ''')
    conn.commit()
    print("[POSTGRESQL] Table created!")

    # Bulk insert
    batch_size = 1000
    total = len(all_data)
    print(f"[POSTGRESQL] Inserting {total} rows...")

    for i in range(0, total, batch_size):
        batch = all_data[i:i+batch_size]
        execute_values(cursor, '''
            INSERT INTO turnover_data (debtor, debtor_name, value, date, date_fixed, branch)
            VALUES %s
        ''', batch)
        conn.commit()
        print(f"  -> Inserted {min(i+batch_size, total)}/{total} rows...")

    print(f"[POSTGRESQL] Done! {total} rows inserted.")
    cursor.close()
    conn.close()

except Exception as e:
    print(f"[POSTGRESQL ERROR] {e}")
    if conn:
        conn.rollback()
        conn.close()

print("\n" + "=" * 50)
print("COMPLETE!")
print("=" * 50)
