"""OneDrive sync module for MOC Automations"""
import os
import json
import io
import re
import time
import threading
import logging
from datetime import datetime
from decimal import Decimal
import msal
import requests
from django.conf import settings
from django.db import connection
from psycopg2.extras import execute_values
import openpyxl
import xlrd

logger = logging.getLogger(__name__)

# Microsoft Graph API endpoint
GRAPH_API_ENDPOINT = 'https://graph.microsoft.com/v1.0'

# Thread lock to prevent concurrent token refresh races
_token_lock = threading.Lock()

# Cached token expiry timestamp (epoch seconds) to avoid unnecessary API calls
_token_expiry = 0


def get_msal_app():
    """Create MSAL application instance"""
    return msal.ConfidentialClientApplication(
        settings.ONEDRIVE_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{settings.ONEDRIVE_TENANT_ID}",
        client_credential=settings.ONEDRIVE_CLIENT_SECRET
    )


def _do_token_refresh(token_data):
    """Attempt to refresh the token using the refresh_token. Returns new access_token or None."""
    global _token_expiry
    refresh_token = token_data.get('refresh_token')
    if not refresh_token:
        logger.error("No refresh_token available - re-authentication required")
        return None

    app = get_msal_app()
    result = app.acquire_token_by_refresh_token(
        refresh_token,
        scopes=settings.ONEDRIVE_SCOPES
    )
    if "access_token" in result:
        save_token(result)
        # Microsoft tokens typically expire in ~3600-5400s, refresh 5 min early
        expires_in = result.get('expires_in', 3600)
        _token_expiry = time.time() + expires_in - 300
        logger.info("Token refreshed successfully")
        return result['access_token']
    else:
        logger.error(f"Token refresh failed: {result.get('error_description', 'Unknown error')}")
        return None


def get_access_token():
    """Get access token from saved file, refresh if expired.

    Uses local expiry tracking to avoid unnecessary API calls.
    Retries refresh up to 3 times with backoff on failure.
    Thread-safe - only one thread refreshes at a time.
    Never returns a known-expired token.
    """
    global _token_expiry

    token_file = settings.ONEDRIVE_TOKEN_FILE
    if not os.path.exists(token_file):
        logger.error("Token file does not exist - run initial OAuth flow")
        return None

    with _token_lock:
        with open(token_file, 'r') as f:
            token_data = json.load(f)

        access_token = token_data.get('access_token')
        if not access_token:
            logger.error("No access_token in token file")
            return None

        # If we know the token hasn't expired yet, return it directly
        if _token_expiry > 0 and time.time() < _token_expiry:
            return access_token

        # Token may be expired - validate with a quick API call
        try:
            headers = {'Authorization': f'Bearer {access_token}'}
            response = requests.get(f'{GRAPH_API_ENDPOINT}/me/drive', headers=headers, timeout=10)

            if response.status_code == 200:
                # Token is valid - set expiry to 45 min from now as a safe estimate
                _token_expiry = time.time() + 2700
                return access_token
        except requests.RequestException as e:
            logger.warning(f"Token validation request failed: {e}")
            # Network issue - if we have a cached expiry that's recent, trust it
            # Otherwise fall through to refresh

        # Token is expired or validation failed - refresh with retries
        logger.info("Access token expired or invalid, refreshing...")
        for attempt in range(1, 4):
            new_token = _do_token_refresh(token_data)
            if new_token:
                return new_token
            if attempt < 3:
                wait = attempt * 5  # 5s, 10s
                logger.warning(f"Token refresh attempt {attempt}/3 failed, retrying in {wait}s...")
                time.sleep(wait)
                # Re-read token file in case another process updated it
                with open(token_file, 'r') as f:
                    token_data = json.load(f)

        logger.error("All 3 token refresh attempts failed")
        return None


def save_token(token_data):
    """Save token data to file atomically"""
    token_file = settings.ONEDRIVE_TOKEN_FILE
    tmp_file = token_file + '.tmp'
    with open(tmp_file, 'w') as f:
        json.dump(token_data, f)
    os.replace(tmp_file, token_file)

def get_auth_url():
    """Generate authorization URL for OAuth flow"""
    app = get_msal_app()
    auth_url = app.get_authorization_request_url(
        scopes=settings.ONEDRIVE_SCOPES,
        redirect_uri=settings.ONEDRIVE_REDIRECT_URI
    )
    return auth_url

def acquire_token_by_auth_code(code):
    """Exchange authorization code for access token"""
    app = get_msal_app()
    result = app.acquire_token_by_authorization_code(
        code,
        scopes=settings.ONEDRIVE_SCOPES,
        redirect_uri=settings.ONEDRIVE_REDIRECT_URI
    )
    if "access_token" in result:
        save_token(result)
        return result
    return None

def list_files_in_folder(folder_path='/'):
    """List all files in OneDrive folder recursively"""
    token = get_access_token()
    if not token:
        return []

    headers = {'Authorization': f'Bearer {token}'}

    # Get folder by path
    if folder_path == '/':
        url = f'{GRAPH_API_ENDPOINT}/me/drive/root/children'
    else:
        url = f'{GRAPH_API_ENDPOINT}/me/drive/root:{folder_path}:/children'

    files = []

    def list_folder(folder_url):
        response = requests.get(folder_url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            for item in data.get('value', []):
                if 'folder' in item:
                    # Recurse into folders
                    list_folder(f"{GRAPH_API_ENDPOINT}/me/drive/items/{item['id']}/children")
                elif 'file' in item:
                    # Add file to list
                    files.append({
                        'id': item['id'],
                        'name': item['name'],
                        'path': item.get('parentReference', {}).get('path', ''),
                        'download_url': item.get('@microsoft.graph.downloadUrl')
                    })

    list_folder(url)
    return files

def download_file(file_id):
    """Download file from OneDrive by file ID"""
    token = get_access_token()
    if not token:
        return None

    headers = {'Authorization': f'Bearer {token}'}
    url = f'{GRAPH_API_ENDPOINT}/me/drive/items/{file_id}/content'

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return io.BytesIO(response.content)
    return None

def delete_file(file_id):
    """Delete file from OneDrive by file ID"""
    token = get_access_token()
    if not token:
        return False

    headers = {'Authorization': f'Bearer {token}'}
    url = f'{GRAPH_API_ENDPOINT}/me/drive/items/{file_id}'

    response = requests.delete(url, headers=headers)
    return response.status_code == 204

def get_branch(filename):
    """Extract branch code from filename"""
    branches = ['ATL', 'HEC', 'HNL', 'HOU', 'ICS', 'IMP', 'JFK', 'LAX', 'LCL', 'ORD', 'PPG']
    filename_upper = filename.upper()

    # Check if filename starts with branch code
    for branch in branches:
        if filename_upper.startswith(branch):
            return branch

    # Check if filename contains branch code
    for branch in branches:
        if branch in filename_upper:
            return branch

    return None

def extract_report_date(ws):
    """Extract report date from row 11 (format: 'Printed by Name DD-MMM-YY HH:MM AM/PM')"""
    try:
        import re
        from datetime import datetime

        row11_text = str(ws.cell(row=11, column=2).value) if hasattr(ws, 'cell') else str(ws.row_values(10)[1])

        # Extract date pattern like "05-Jan-26 12:06 AM"
        date_match = re.search(r'(\d{2}-[A-Za-z]{3}-\d{2})\s+(\d{1,2}:\d{2}\s+[AP]M)', row11_text)
        if date_match:
            date_str = date_match.group(1)
            time_str = date_match.group(2)
            report_date = datetime.strptime(f"{date_str} {time_str}", "%d-%b-%y %I:%M %p")
            return report_date.strftime('%Y-%m-%d')
    except:
        pass
    return None

KNOWN_BRANCHES = ['ATL', 'HEC', 'HNL', 'HOU', 'ICS', 'IMP', 'JFK', 'LAX', 'LCL', 'ORD', 'PPG', 'CON', 'DOR']

def get_branch_from_file(ws):
    """Extract branch code from inside the Excel file by scanning header rows."""
    for row_idx in range(1, 12):
        try:
            for col_idx in range(1, 5):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                # Match "Transaction Branches: IMP" style
                m = re.search(r'Transaction Branches?:\s*([A-Z\-]+)', cell_val, re.IGNORECASE)
                if m:
                    candidate = m.group(1).strip().upper()
                    if candidate in KNOWN_BRANCHES:
                        return candidate
        except:
            continue
    return None


def process_excel_file(file_content, filename):
    """Process Excel file and return data rows with report date"""
    branch = get_branch(filename)

    rows = []

    try:
        # Try openpyxl for .xlsx
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb.active

        # If filename didn't reveal branch, read it from inside the file
        if not branch:
            branch = get_branch_from_file(ws)
        if not branch:
            print(f"  Could not determine branch for {filename} - skipping")
            return []

        # Extract report date from row 11
        report_date = extract_report_date(ws)

        # Check if this is WIDE format (header row with month columns)
        # Row 13 should have: Debtor, Debtor Name, TOTAL, 202501, 202502, etc.
        header_row = [ws.cell(row=13, column=i).value for i in range(4, 20)]

        # Check if we have month columns (202501, 202502, etc.)
        month_columns = []
        for col_idx, val in enumerate(header_row, start=4):
            if isinstance(val, int) and 202000 <= val <= 209912:  # Valid YYYYMM
                month_columns.append((col_idx, val))

        if month_columns:
            # WIDE FORMAT: Pivot data from columns to rows
            print(f"  Processing WIDE format with {len(month_columns)} month columns")

            # Detect debtor and debtor_name columns from header row 13
            debtor_col, name_col = 4, 5
            full_header = [ws.cell(row=13, column=i).value for i in range(1, ws.max_column + 1)]
            for ci, v in enumerate(full_header, start=1):
                if v and str(v).strip().lower() == 'debtor':
                    debtor_col = ci
                if v and str(v).strip().lower() in ('debtor name', 'name'):
                    name_col = ci

            # Process data rows (starting from row 14)
            for row_idx in range(14, ws.max_row + 1):
                debtor = ws.cell(row=row_idx, column=debtor_col).value
                if not debtor:
                    continue

                debtor = str(debtor).strip()
                debtor_name = ws.cell(row=row_idx, column=name_col).value
                debtor_name = str(debtor_name).strip() if debtor_name else ''

                # Process each month column
                for col_idx, month_val in month_columns:
                    value = ws.cell(row=row_idx, column=col_idx).value

                    if value and isinstance(value, (int, float)) and value != 0:
                        # Convert month to date string
                        year = month_val // 100
                        month = month_val % 100
                        date_str = f"{year:04d}-{month:02d}-01"

                        rows.append((debtor, debtor_name, date_str, branch, value, report_date))

        else:
            # LONG FORMAT: Original processing (one row per value)
            print(f"  Processing LONG format")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4 and row[0]:
                    debtor = str(row[0]).strip() if row[0] else None
                    debtor_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    date_val = row[2] if len(row) > 2 else None
                    value = float(row[3]) if len(row) > 3 and row[3] else 0

                    if not debtor or not date_val:
                        continue

                    # Parse date
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime('%Y-%m-%d')
                    elif isinstance(date_val, (int, float)):
                        date_int = int(date_val)
                        year = date_int // 100
                        month = date_int % 100
                        date_str = f"{year:04d}-{month:02d}-01"
                    else:
                        try:
                            date_int = int(str(date_val).strip())
                            year = date_int // 100
                            month = date_int % 100
                            date_str = f"{year:04d}-{month:02d}-01"
                        except:
                            date_str = str(date_val) if date_val else None

                    if debtor and date_str and value != 0:
                        rows.append((debtor, debtor_name, date_str, branch, value, report_date))

    except Exception as e:
        print(f"  Error with openpyxl: {e}, trying xlrd...")
        # Try xlrd for .xls
        try:
            file_content.seek(0)
            wb = xlrd.open_workbook(file_contents=file_content.read())
            ws = wb.sheet_by_index(0)

            report_date = extract_report_date(ws)

            # Check for wide format
            header_row = ws.row_values(12)  # Row 13 (0-indexed as 12)
            month_columns = []
            for col_idx, val in enumerate(header_row[3:19], start=3):
                if isinstance(val, (int, float)) and 202000 <= val <= 209912:
                    month_columns.append((col_idx, int(val)))

            if month_columns:
                # WIDE FORMAT
                # Detect debtor and name columns from header row (0-indexed row 12)
                hdr = ws.row_values(12)
                debtor_col, name_col = 3, 4
                for ci, v in enumerate(hdr):
                    if v and str(v).strip().lower() == 'debtor':
                        debtor_col = ci
                    if v and str(v).strip().lower() in ('debtor name', 'name'):
                        name_col = ci

                for row_idx in range(13, ws.nrows):  # Row 14+ (0-indexed as 13+)
                    row = ws.row_values(row_idx)
                    debtor = row[debtor_col] if len(row) > debtor_col else None
                    if not debtor:
                        continue

                    debtor = str(debtor).strip()
                    debtor_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ''

                    for col_idx, month_val in month_columns:
                        value = row[col_idx] if len(row) > col_idx else None
                        if value and isinstance(value, (int, float)) and value != 0:
                            year = month_val // 100
                            month = month_val % 100
                            date_str = f"{year:04d}-{month:02d}-01"
                            rows.append((debtor, debtor_name, date_str, branch, value, report_date))
            else:
                # LONG FORMAT
                for row_idx in range(1, ws.nrows):
                    row = ws.row_values(row_idx)
                    if row and len(row) >= 4 and row[0]:
                        debtor = str(row[0]).strip() if row[0] else None
                        debtor_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                        date_val = row[2] if len(row) > 2 else None
                        value = float(row[3]) if len(row) > 3 and row[3] else 0

                        if not debtor or not date_val:
                            continue

                        if isinstance(date_val, (int, float)):
                            date_int = int(date_val)
                            year = date_int // 100
                            month = date_int % 100
                            date_str = f"{year:04d}-{month:02d}-01"
                        else:
                            try:
                                date_int = int(str(date_val).strip())
                                year = date_int // 100
                                month = date_int % 100
                                date_str = f"{year:04d}-{month:02d}-01"
                            except:
                                date_str = str(date_val)

                        if debtor and date_str and value != 0:
                            rows.append((debtor, debtor_name, date_str, branch, value, report_date))
        except Exception as e:
            print(f"  Error processing {filename}: {e}")
            return []

    return rows

def sync_turnover_data():
    """Sync turnover data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files, skip historical year archives (20XX_*.xlsx)
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
        and not re.match(r'^20\d{2}_', f['name'])  # Skip 2020_ATL.xlsx, 2021_HNL.xlsx, etc.
    ]

    all_rows = []

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")

    # Upsert to database with report date tracking
    if all_rows:
        with connection.cursor() as cur:
            # New query that only updates if the new report is newer (or if no report_date exists)
            query = """
                INSERT INTO turnover_data (debtor, debtor_name, date, branch, value, report_date)
                VALUES %s
                ON CONFLICT (debtor, date, branch)
                DO UPDATE SET
                    value = EXCLUDED.value,
                    debtor_name = EXCLUDED.debtor_name,
                    report_date = EXCLUDED.report_date
                WHERE
                    turnover_data.report_date IS NULL
                    OR EXCLUDED.report_date IS NULL
                    OR EXCLUDED.report_date >= turnover_data.report_date
            """
            execute_values(cur, query, all_rows)
        print(f"\n✓ Total: {len(all_rows)} records processed")
    else:
        print(f"\n✓ No new files to sync")

    # Always save last sync time (even if no new records) - using local system time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def yyyymm_to_date(month_yyyymm):
    """Convert YYYYMM integer to date string (first day of month).
    Example: 202601 -> '2026-01-01'
    """
    year = month_yyyymm // 100
    month = month_yyyymm % 100
    return f"{year}-{month:02d}-01"


def calculate_week_number(report_date, month_yyyymm):
    """Calculate week number based on report date and data month.
    Week 1 = 8th, Week 2 = 15th, Week 3 = 22nd, Week 4 = 29th, Week 5 = end of month.
    Feb only has 4 weeks: Week 1=8th, Week 2=15th, Week 3=22nd, Week 4=28th (final).
    For previous months (past month end), Week 5 (or Week 4 for Feb).
    """
    # Extract year and month from YYYYMM
    data_year = month_yyyymm // 100
    data_month = month_yyyymm % 100

    # Get the month/year from report_date
    report_year = report_date.year
    report_month = report_date.month

    # If the data month is BEFORE the report month, it's past the month end
    # Feb only has 4 weeks, so final week = Week 4. All other months = Week 5.
    if (data_year < report_year) or (data_year == report_year and data_month < report_month):
        if data_month == 2:
            return 4  # Feb final week is Week 4
        return 5  # Final week for previous month

    # Same month as report - calculate based on report date day
    day = report_date.day

    # Days 1-7 = No week (too early!)
    if day <= 7:
        return None  # Too early - no week yet

    # February: only 4 weeks (no 29th), Week 1=8, Week 2=15, Week 3=22, Week 4=28 (final)
    if data_month == 2:
        if day <= 14:
            return 1
        elif day <= 21:
            return 2
        elif day <= 27:
            return 3
        else:
            return 4

    # All other months: fixed 7-day intervals
    # Week 1: 8-14, Week 2: 15-21, Week 3: 22-28, Week 4: 29+
    if day <= 14:
        return 1
    elif day <= 21:
        return 2
    elif day <= 28:
        return 3
    else:
        return 4


def process_ppg_excel_file(file_content, filename):
    """Process PPG Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = []

    try:
        # Extract report date from filename
        # Format: "PPG INT Profit And Loss Period Analysis 202602 Sunday, 01 February 2026 00_04_39.XLS"
        import re
        date_match = re.search(r'(\d{2})\s+(\w+)\s+(\d{4})', filename)
        if date_match:
            day = int(date_match.group(1))
            month_name = date_match.group(2)
            year = int(date_match.group(3))
            report_date = datetime.strptime(f"{day} {month_name} {year}", "%d %B %Y")
            print(f"  Report date: {report_date.strftime('%Y-%m-%d')}")
        else:
            print(f"  Warning: Could not extract report date from filename")
            report_date = datetime.now()

        file_content.seek(0)

        # Try openpyxl first for .xlsx files
        try:
            wb = openpyxl.load_workbook(file_content, data_only=True)
            ws = wb['GL PL Period Analysis']

            print(f"  Processing sheet: GL PL Period Analysis")
            print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")

            # Find header row (row with month columns like 202601, 202602)
            header_row_idx = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                # Check cells in this row for YYYYMM pattern
                for col in range(1, min(30, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col).value
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        header_row_idx = row_idx
                        break
                if header_row_idx:
                    break

            if header_row_idx is None:
                print(f"  Error: Could not find header row with month columns")
                return []

            print(f"  Found header row at row {header_row_idx}")

            # Find month columns by scanning the header row
            # Only keep FIRST occurrence of each month (Excel may have duplicate month columns)
            month_columns = []
            seen_months = set()
            for col_num in range(1, min(60, ws.max_column + 1)):  # Scan more columns
                val = ws.cell(row=header_row_idx, column=col_num).value
                if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                    month_val = int(val)
                    if month_val not in seen_months:
                        month_columns.append((col_num, month_val))
                        seen_months.add(month_val)

            print(f"  Found {len(month_columns)} month columns: {month_columns}")

            # Process data rows
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                # Don't use list comprehension - directly access cells by column index
                # to avoid issues with variable row lengths
                if not any(ws[row_idx]):
                    continue

                # Column 2: A/C (account number), Column 3: Account name (Excel is 1-indexed)
                account_number = str(ws.cell(row=row_idx, column=2).value).strip() if ws.cell(row=row_idx, column=2).value else None
                account_name = str(ws.cell(row=row_idx, column=3).value).strip() if ws.cell(row=row_idx, column=3).value else None

                if not account_number and not account_name:
                    continue

                # Debug logging for specific account
                debug_account = "CURRENT YEAR INCOME (LOSS)"
                if account_name and debug_account in account_name:
                    print(f"  DEBUG: Row {row_idx}, Account: {account_name}")

                # Process each month column
                for col_num, month_yyyymm in month_columns:
                    # Use direct cell access instead of row list to avoid index misalignment
                    # col_num is now 1-indexed (Excel column number)
                    value = ws.cell(row=row_idx, column=col_num).value

                    # Debug logging
                    if account_name and debug_account in account_name:
                        print(f"    Col {col_num} (month {month_yyyymm}): value = {value}")

                    if value is not None and isinstance(value, (int, float)):
                        # Calculate week number for this row
                        week = calculate_week_number(report_date, month_yyyymm)

                        # Skip if no week assigned (too early in the month)
                        if week is None:
                            continue

                        week_label = f"Week {week}"

                        # Convert YYYYMM to date
                        date_fixed = yyyymm_to_date(month_yyyymm)

                        # OneDrive sync only handles Actual data - Budget stays completely untouched
                        budget_actual = "Actual"

                        # Store: (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                        rows.append(("PPG", account_name, value, month_yyyymm, date_fixed, budget_actual, week_label, report_date.strftime('%Y-%m-%d')))

        except Exception as xlsx_error:
            # Try xlrd for old .xls format
            file_content.seek(0)
            wb = xlrd.open_workbook(file_contents=file_content.read())
            ws = wb.sheet_by_name('GL PL Period Analysis')

            print(f"  Processing sheet: GL PL Period Analysis")
            print(f"  Rows: {ws.nrows}, Columns: {ws.ncols}")

            # Find header row (row with month columns like 202601, 202602)
            header_row_idx = None
            for row_idx in range(min(20, ws.nrows)):
                # Check cells in this row for YYYYMM pattern
                for col_idx in range(min(30, ws.ncols)):
                    val = ws.cell_value(row_idx, col_idx)
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        header_row_idx = row_idx
                        break
                if header_row_idx is not None:
                    break

            if header_row_idx is None:
                print(f"  Error: Could not find header row with month columns")
                return []

            print(f"  Found header row at row {header_row_idx + 1}")

            # Find month columns by scanning the header row
            # Only keep FIRST occurrence of each month (Excel may have duplicate month columns)
            month_columns = []
            seen_months = set()
            for col_idx in range(min(60, ws.ncols)):  # Scan more columns
                val = ws.cell_value(header_row_idx, col_idx)
                if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                    month_val = int(val)
                    if month_val not in seen_months:
                        month_columns.append((col_idx, month_val))
                        seen_months.add(month_val)

            print(f"  Found {len(month_columns)} month columns: {month_columns}")

            # Process data rows
            for row_idx in range(header_row_idx + 1, ws.nrows):
                # Check if row has any data
                row = ws.row_values(row_idx)
                if not any(row):
                    continue

                # Column 1: A/C (account number), Column 2: Account name (0-indexed)
                account_number = str(ws.cell_value(row_idx, 1)).strip() if ws.ncols > 1 and ws.cell_value(row_idx, 1) else None
                account_name = str(ws.cell_value(row_idx, 2)).strip() if ws.ncols > 2 and ws.cell_value(row_idx, 2) else None

                if not account_number and not account_name:
                    continue

                # Debug logging for specific account
                debug_account = "CURRENT YEAR INCOME (LOSS)"
                if account_name and debug_account in account_name:
                    print(f"  DEBUG: Row {row_idx + 1}, Account: {account_name}")

                # Process each month column
                for col_idx, month_yyyymm in month_columns:
                    if col_idx >= ws.ncols:
                        continue

                    # Use direct cell access instead of row list to avoid index misalignment
                    value = ws.cell_value(row_idx, col_idx)

                    # Debug logging
                    if account_name and debug_account in account_name:
                        print(f"    Col {col_idx} (month {month_yyyymm}): value = {value}")

                    if value is not None and isinstance(value, (int, float)):
                        # Calculate week number for this row
                        week = calculate_week_number(report_date, month_yyyymm)

                        # Skip if no week assigned (too early in the month)
                        if week is None:
                            continue

                        week_label = f"Week {week}"

                        # Convert YYYYMM to date
                        date_fixed = yyyymm_to_date(month_yyyymm)

                        # OneDrive sync only handles Actual data - Budget stays completely untouched
                        budget_actual = "Actual"

                        # Store: (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                        rows.append(("PPG", account_name, value, month_yyyymm, date_fixed, budget_actual, week_label, report_date.strftime('%Y-%m-%d')))

    except Exception as e:
        print(f"  Error processing PPG file {filename}: {e}")
        import traceback
        traceback.print_exc()
        return []

    return rows


def sync_ppg_data():
    """Sync PPG data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_PPG_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in PPG folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_ppg_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ppg_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM ppg_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM ppg_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO ppg_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ppg_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time (using local system time)
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.PPG_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_ppg_last_sync():
    """Get the last PPG sync info"""
    try:
        with open(settings.PPG_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


def process_dor_excel_file(file_content, filename):
    """Process DOR Excel file - reads 'GL PL Period Analysis' sheet"""
    # Same logic as PPG
    return process_ppg_excel_file(file_content, filename.replace('DOR', 'PPG'))


def sync_dor_data():
    """Sync DOR data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_DOR_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in DOR folder")

    all_rows = []
    processed_files = []

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_dor_excel_file(file_content, file_info['name'])
            # Change division from PPG to DOR
            rows = [(('DOR', row[1], row[2], row[3], row[4], row[5], row[6], row[7])) for row in rows]
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)

    # Insert data
    if all_rows:
        with connection.cursor() as cur:
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM dor_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM dor_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM dor_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Deduplicate rows (DOR has 3 sub-files that may overlap on same account)
            # Key = (division, account_name, date, budget_actual, week), keep last value
            deduped = {}
            for row in all_rows:
                key = (row[0], row[1], row[3], row[5], row[6])  # div, acct, date, budget_actual, week
                deduped[key] = row
            all_rows = list(deduped.values())

            # Insert new week data
            week_query = """
                INSERT INTO dor_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM dor_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.DOR_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_dor_last_sync():
    """Get the last DOR sync info"""
    try:
        with open(settings.DOR_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

def process_con_excel_file(file_content, filename):
    """Process CON Excel file - reads 'GL PL Period Analysis' sheet"""
    # Same logic as PPG
    return process_ppg_excel_file(file_content, filename.replace('CON', 'PPG'))


def process_atl_excel_file(file_content, filename):
    """Process ATL Excel file - reads 'GL PL Period Analysis' sheet"""
    # ATL uses IMP as the division, so we need to process like PPG then replace division
    rows = process_ppg_excel_file(file_content, filename.replace('ATL', 'PPG'))
    # Replace PPG division with IMP (ATL's division code)
    return [(("IMP",) + row[1:]) for row in rows]


def process_hnl_excel_file(file_content, filename):
    """Process HNL Excel file - reads 'GL PL Period Analysis' sheet"""
    # HNL uses HNL as the division, just process like PPG and keep HNL
    rows = process_ppg_excel_file(file_content, filename.replace('HNL', 'PPG'))
    # Replace PPG division with HNL
    return [(("HNL",) + row[1:]) for row in rows]


def sync_con_data():
    """Sync CON data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_CON_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in CON folder")

    all_rows = []
    processed_files = []

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_con_excel_file(file_content, file_info['name'])
            # Change division from PPG to CON
            rows = [(('CON', row[1], row[2], row[3], row[4], row[5], row[6], row[7])) for row in rows]
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)

    # Insert data
    if all_rows:
        with connection.cursor() as cur:
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM con_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM con_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM con_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO con_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM con_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.CON_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_con_last_sync():
    """Get the last CON sync info"""
    try:
        with open(settings.CON_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


def sync_atl_data():
    """Sync ATL data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_ATL_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in ATL folder")

    all_rows = []
    processed_files = []

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            # Process Excel file using ATL processor (similar to PPG/CON)
            rows = process_atl_excel_file(file_content, file_info['name'])

            if rows:
                all_rows.extend(rows)
                print(f"  Extracted {len(rows)} records")
                processed_files.append(file_info)
            else:
                print(f"  No data extracted from file")

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM atl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM atl_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM atl_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO atl_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM atl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.ATL_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_atl_last_sync():
    """Get the last ATL sync info"""
    try:
        with open(settings.ATL_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


def sync_hnl_data():
    """Sync HNL data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_HNL_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in HNL folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_hnl_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM hnl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM hnl_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM hnl_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO hnl_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM hnl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.HNL_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_hnl_last_sync():
    """Get the last HNL sync info"""
    try:
        with open(settings.HNL_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


# ==================== JFK Functions ====================

def process_jfk_excel_file(file_content, filename):
    """Process JFK Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('JFK', 'PPG'))
    return [(("JFK",) + row[1:]) for row in rows]


def sync_jfk_data():
    """Sync JFK data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_JFK_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in JFK folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_jfk_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM jfk_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM jfk_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM jfk_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO jfk_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM jfk_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.JFK_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_jfk_last_sync():
    """Get the last JFK sync info"""
    try:
        with open(settings.JFK_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


# ==================== CCC Functions ====================

def process_ccc_excel_file(file_content, filename):
    """Process CCC Excel file - reads 'GL PL Period Analysis' sheet and adds CCC division"""
    # First process using PPG logic
    rows = process_ppg_excel_file(file_content, filename.replace('CCC', 'PPG'))

    # Replace PPG with CCC in division column
    return [(("CCC",) + row[1:]) for row in rows]


def sync_ccc_data():
    """Sync CCC data from OneDrive to PostgreSQL - only updates specific weeks, preserves historical data"""
    print("\n" + "=" * 60)
    print("Starting CCC OneDrive Sync")
    print("=" * 60)

    # Get access token
    token = get_access_token()
    if not token:
        raise Exception("Failed to get access token")

    # Get all Excel files from OneDrive folder
    folder_path = settings.ONEDRIVE_CCC_FOLDER_PATH
    headers = {'Authorization': f'Bearer {token}'}

    # Get folder ID
    folder_url = f"https://graph.microsoft.com/v1.0/me/drive/root:{folder_path}"
    folder_response = requests.get(folder_url, headers=headers)

    if folder_response.status_code != 200:
        raise Exception(f"Failed to access folder: {folder_response.text}")

    folder_id = folder_response.json()['id']

    # List all files in the folder
    files_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{folder_id}/children"
    files_response = requests.get(files_url, headers=headers)

    if files_response.status_code != 200:
        raise Exception(f"Failed to list files: {files_response.text}")

    files = files_response.json().get('value', [])

    # Filter for Excel files (case insensitive extension check)
    excel_files = [f for f in files if f['name'].lower().endswith('.xlsx') or f['name'].lower().endswith('.xls')]

    if not excel_files:
        print("No Excel files found in CCC folder")
        from zoneinfo import ZoneInfo
        local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
        with open(settings.CCC_LAST_SYNC_FILE, 'w') as f:
            json.dump({'last_sync': local_time.isoformat(), 'had_data': False}, f)
        return 0

    print(f"\nFound {len(excel_files)} Excel file(s) with '2026' in name:")
    for f in excel_files:
        print(f"  - {f['name']}")

    # Process all files
    all_rows = []
    processed_files = []  # Track successfully processed files
    for file in excel_files:
        print(f"\nProcessing: {file['name']}")

        # Download file
        download_url = file['@microsoft.graph.downloadUrl']
        file_response = requests.get(download_url)

        if file_response.status_code != 200:
            print(f"  ✗ Failed to download")
            continue

        # Process file
        try:
            rows = process_ccc_excel_file(io.BytesIO(file_response.content), file['name'])
            print(f"  ✓ Processed {len(rows)} rows")
            all_rows.extend(rows)
            processed_files.append(file)  # Track for deletion
        except Exception as e:
            print(f"  ✗ Error processing file: {str(e)}")
            continue

    if not all_rows:
        print("\n✓ No data to sync")
        from zoneinfo import ZoneInfo
        local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
        with open(settings.CCC_LAST_SYNC_FILE, 'w') as f:
            json.dump({'last_sync': local_time.isoformat(), 'had_data': False}, f)
        return 0

    print(f"\nTotal rows from all files: {len(all_rows)}")

    # Connect to database using Django connection
    from django.db import connection
    with connection.cursor() as cur:
        # Get unique (month, week) combinations from the new data
        unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
        unique_months = set(row[3] for row in all_rows)

        # DEBUG: Show existing weeks before deletion
        print(f"\n🔍 DEBUG - Before deletion:")
        for month in sorted(unique_months):
            cur.execute("""
                SELECT week, COUNT(*)
                FROM ccc_pnl
                WHERE date = %s AND budget_actual = 'Actual'
                GROUP BY week
                ORDER BY week
            """, (month,))
            weeks = cur.fetchall()
            if weeks:
                week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                print(f"  Month {month}: {week_summary}")

        # Show what will be deleted
        print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
        for month, week in sorted(unique_month_weeks):
            print(f"  - {month}, {week}")

        # Delete old Actual week records only (Budget stays untouched)
        # Only delete the specific (month, week) combinations that we're updating
        for month, week in unique_month_weeks:
            cur.execute(
                "DELETE FROM ccc_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                (month, week)
            )
        print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

        # Delete old Actual "Total" records only (Budget stays untouched)
        for month in unique_months:
            cur.execute(
                "DELETE FROM ccc_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                (month,)
            )
        print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

        # Insert new week data
        week_query = """
            INSERT INTO ccc_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
            VALUES %s
        """
        execute_values(cur, week_query, all_rows)
        print(f"  Inserted {len(all_rows)} week records")

        # Duplicate rows as "Total" - separately for each month to keep them separate
        # Group rows by month
        from collections import defaultdict
        rows_by_month = defaultdict(list)
        for row in all_rows:
            month = row[3]  # date (YYYYMM)
            rows_by_month[month].append(row)

        # For each month, duplicate its rows as Total
        total_count = 0
        for month, month_rows in rows_by_month.items():
            total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
            execute_values(cur, week_query, total_rows)
            total_count += len(total_rows)
            print(f"  Inserted {len(total_rows)} Total records for month {month}")

        print(f"  Total: {total_count} Total records inserted")

        # DEBUG: Show weeks after insertion
        print(f"\n🔍 DEBUG - After insertion:")
        for month in sorted(unique_months):
            cur.execute("""
                SELECT week, COUNT(*)
                FROM ccc_pnl
                WHERE date = %s AND budget_actual = 'Actual'
                GROUP BY week
                ORDER BY week
            """, (month,))
            weeks = cur.fetchall()
            if weeks:
                week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                print(f"  Month {month}: {week_summary}")

        print("\n" + "=" * 60)
        print("✓ CCC SYNC COMPLETED!")
        print("=" * 60)

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file in processed_files:
            delete_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file['id']}"
            delete_response = requests.delete(delete_url, headers=headers)
            if delete_response.status_code == 204:
                print(f"  ✓ Deleted: {file['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    with open(settings.CCC_LAST_SYNC_FILE, 'w') as f:
        json.dump({'last_sync': local_time.isoformat(), 'had_data': True}, f)

    return len(all_rows)


def get_ccc_last_sync():
    """Get the last CCC sync info"""
    try:
        with open(settings.CCC_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


# ==================== CCD Functions ====================

def process_ccd_excel_file(file_content, filename):
    """Process CCD Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('CCD', 'PPG'))
    return [(("CCD",) + row[1:]) for row in rows]


def sync_ccd_data():
    """Sync CCD data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_CCD_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in CCD folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_ccd_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ccd_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM ccd_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM ccd_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO ccd_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ccd_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.CCD_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_ccd_last_sync():
    """Get the last CCD sync info"""
    try:
        with open(settings.CCD_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== FAX Functions ====================

def process_fax_excel_file(file_content, filename):
    """Process FAX Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('FAX', 'PPG'))
    return [(("FAX",) + row[1:]) for row in rows]


def sync_fax_data():
    """Sync FAX data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_FAX_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in FAX folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_fax_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM fax_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM fax_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM fax_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO fax_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM fax_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.FAX_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_fax_last_sync():
    """Get the last FAX sync info"""
    try:
        with open(settings.FAX_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== IMP Functions ====================

def process_imp_excel_file(file_content, filename):
    """Process IMP Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('IMP', 'PPG'))
    return [(("IMP",) + row[1:]) for row in rows]


def sync_imp_data():
    """Sync IMP data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_IMP_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in IMP folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_imp_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM imp_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM imp_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM imp_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO imp_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM imp_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.IMP_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_imp_last_sync():
    """Get the last IMP sync info"""
    try:
        with open(settings.IMP_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== HOU Functions ====================

def process_hou_excel_file(file_content, filename):
    """Process HOU Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('HOU', 'PPG'))
    return [(("HOU",) + row[1:]) for row in rows]


def sync_hou_data():
    """Sync HOU data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_HOU_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in HOU folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_hou_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM hou_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM hou_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM hou_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO hou_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM hou_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.HOU_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_hou_last_sync():
    """Get the last HOU sync info"""
    try:
        with open(settings.HOU_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== ICS Functions ====================

def process_ics_excel_file(file_content, filename):
    """Process ICS Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('ICS', 'PPG'))
    return [(("ICS",) + row[1:]) for row in rows]


def sync_ics_data():
    """Sync ICS data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_ICS_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in ICS folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_ics_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ics_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM ics_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM ics_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO ics_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ics_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.ICS_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_ics_last_sync():
    """Get the last ICS sync info"""
    try:
        with open(settings.ICS_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== LAX Functions ====================

def process_lax_excel_file(file_content, filename):
    """Process LAX Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('LAX', 'PPG'))
    return [(("LAX",) + row[1:]) for row in rows]


def sync_lax_data():
    """Sync LAX data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_LAX_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in LAX folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_lax_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM lax_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM lax_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM lax_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO lax_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM lax_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.LAX_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_lax_last_sync():
    """Get the last LAX sync info"""
    try:
        with open(settings.LAX_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== LCL Functions ====================

def process_lcl_excel_file(file_content, filename):
    """Process LCL Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('LCL', 'PPG'))
    return [(("LCL",) + row[1:]) for row in rows]


def sync_lcl_data():
    """Sync LCL data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_LCL_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in LCL folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_lcl_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM lcl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM lcl_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM lcl_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO lcl_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM lcl_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.LCL_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_lcl_last_sync():
    """Get the last LCL sync info"""
    try:
        with open(settings.LCL_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

# ==================== ORD Functions ====================

def process_ord_excel_file(file_content, filename):
    """Process ORD Excel file - reads 'GL PL Period Analysis' sheet"""
    rows = process_ppg_excel_file(file_content, filename.replace('ORD', 'PPG'))
    return [(("ORD",) + row[1:]) for row in rows]


def sync_ord_data():
    """Sync ORD data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_ORD_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in ORD folder")

    all_rows = []
    processed_files = []  # Track successfully processed files

    for file_info in excel_files:
        print(f"Processing {file_info['name']}...")
        file_content = download_file(file_info['id'])

        if file_content:
            rows = process_ord_excel_file(file_content, file_info['name'])
            all_rows.extend(rows)
            print(f"  Extracted {len(rows)} records")
            processed_files.append(file_info)  # Track for deletion

    # Insert data with proper week and Total handling
    if all_rows:
        with connection.cursor() as cur:
            # Get unique (month, week) combinations from the new data
            unique_month_weeks = set((row[3], row[6]) for row in all_rows)  # (month, week)
            unique_months = set(row[3] for row in all_rows)

            # DEBUG: Show existing weeks before deletion
            print(f"\n🔍 DEBUG - Before deletion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ord_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

            # Show what will be deleted
            print(f"\n🗑️  Will DELETE and replace these (month, week) combinations:")
            for month, week in sorted(unique_month_weeks):
                print(f"  - {month}, {week}")

            # Delete old Actual week records only (Budget stays untouched)
            for month, week in unique_month_weeks:
                cur.execute(
                    "DELETE FROM ord_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'",
                    (month, week)
                )
            print(f"\n✓ Deleted old Actual week records for {len(unique_month_weeks)} (month, week) combinations")

            # Delete old Actual "Total" records only (Budget stays untouched)
            for month in unique_months:
                cur.execute(
                    "DELETE FROM ord_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'",
                    (month,)
                )
            print(f"  Deleted old Actual 'Total' records for {len(unique_months)} months")

            # Insert new week data
            week_query = """
                INSERT INTO ord_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date)
                VALUES %s
            """
            execute_values(cur, week_query, all_rows)
            print(f"  Inserted {len(all_rows)} week records")

            # Duplicate rows as "Total" - separately for each month to keep them separate
            # Group rows by month
            from collections import defaultdict
            rows_by_month = defaultdict(list)
            for row in all_rows:
                month = row[3]  # date (YYYYMM)
                rows_by_month[month].append(row)

            # For each month, duplicate its rows as Total
            total_count = 0
            for month, month_rows in rows_by_month.items():
                total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
                execute_values(cur, week_query, total_rows)
                total_count += len(total_rows)
                print(f"  Inserted {len(total_rows)} Total records for month {month}")

            print(f"  Total: {total_count} Total records inserted")

            # DEBUG: Show existing weeks after insertion
            print(f"\n🔍 DEBUG - After insertion:")
            for month in sorted(unique_months):
                cur.execute("""
                    SELECT week, COUNT(*)
                    FROM ord_pnl
                    WHERE date = %s AND budget_actual = 'Actual'
                    GROUP BY week
                    ORDER BY week
                """, (month,))
                weeks = cur.fetchall()
                if weeks:
                    week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
                    print(f"  Month {month}: {week_summary}")

        print(f"\n✓ Imported {len(all_rows)} week records + {len(all_rows)} Total records")
        print(f"✅ Other weeks preserved (not deleted)")

        # Delete processed files from OneDrive
        print(f"\nDeleting {len(processed_files)} processed files from OneDrive...")
        deleted_count = 0
        for file_info in processed_files:
            if delete_file(file_info['id']):
                print(f"  ✓ Deleted: {file_info['name']}")
                deleted_count += 1
            else:
                print(f"  ✗ Failed to delete: {file_info['name']}")
        print(f"✓ Deleted {deleted_count}/{len(processed_files)} files from OneDrive")
    else:
        print(f"\n✓ No files to sync")

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    last_sync = {'last_sync': local_time.isoformat(), 'had_data': len(all_rows) > 0}
    with open(settings.ORD_LAST_SYNC_FILE, 'w') as f:
        json.dump(last_sync, f)

    return len(all_rows)


def get_ord_last_sync():
    """Get the last ORD sync info"""
    try:
        with open(settings.ORD_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None


def sync_creditor_data():
    """Sync all creditor group files from OneDrive to PostgreSQL"""
    groups = ['AGT', 'AIR', 'ASC', 'ATT', 'AUS', 'CRE', 'EMP', 'HR', 'INT', 'OH', 'SSL', 'TPY', 'TRU', 'X2']
    grand_total = 0

    for group in groups:
        folder_path = f'/Automation Platform/Creditor Report/{group}'
        files = list_files_in_folder(folder_path)
        excel_files = [f for f in files if f['name'].lower().endswith(('.xlsx', '.xls'))]

        if not excel_files:
            print(f'{group}: No file')
            continue

        # Prefer 'Creditor Transaction Report' files over other files
        ctr_files = [f for f in excel_files if 'creditor transaction report' in f['name'].lower()]
        file_info = ctr_files[0] if ctr_files else excel_files[0]
        print(f'{group}: Downloading {file_info["name"]}')
        content = download_file(file_info['id'])

        if not content:
            print(f'{group}: Download failed')
            continue

        raw_bytes = content.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column
            def cell(r, c, _ws=ws):
                return _ws.cell(row=r, column=c).value
        except Exception:
            try:
                wbx = xlrd.open_workbook(file_contents=raw_bytes)
                wsx = wbx.sheet_by_index(0)
                max_row = wsx.nrows
                max_col = wsx.ncols
                def cell(r, c, _wsx=wsx):
                    if r-1 < _wsx.nrows and c-1 < _wsx.ncols:
                        v = _wsx.cell_value(r-1, c-1)
                        return v if v != '' else None
                    return None
            except Exception as e:
                print(f'{group}: Error opening file - {e}')
                continue

        # Get period columns from row 8
        periods = []
        for col_idx in range(5, max_col + 1):
            v = cell(8, col_idx)
            if v is not None:
                vs = str(v).strip().split('.')[0]
                if re.match(r'^\d{6}$', vs):
                    periods.append((col_idx, vs))

        if not periods:
            print(f'{group}: No period columns')
            continue

        current_branch = ''
        all_rows = []
        for row_idx in range(9, max_row + 1):
            col_b = cell(row_idx, 2)
            if col_b is None:
                continue
            col_b = str(col_b).strip()
            if col_b.startswith('Organization Branch:'):
                match = re.search(r'\(([^)]*)\)', col_b)
                current_branch = match.group(1) if match else ''
                continue
            if 'Total' in col_b or 'Grand' in col_b or col_b == '':
                continue
            creditor = col_b
            creditor_name = str(cell(row_idx, 3) or '').strip()
            if not creditor_name:
                continue
            for col_idx, period in periods:
                val = cell(row_idx, col_idx)
                if val is not None and str(val).strip() != '':
                    try:
                        from decimal import Decimal
                        value = Decimal(str(val).replace(',', ''))
                    except Exception:
                        continue
                    all_rows.append((creditor, creditor_name, period, value, group, current_branch))

        if all_rows:
            periods_in_data = sorted(set(r[2] for r in all_rows))
            with connection.cursor() as cursor:
                for period in periods_in_data:
                    cursor.execute(
                        "DELETE FROM creditor_transactions WHERE creditor_group = %s AND period = %s",
                        (group, period)
                    )
                execute_values(
                    cursor,
                    "INSERT INTO creditor_transactions (creditor, creditor_name, period, value, creditor_group, branch) VALUES %s",
                    all_rows
                )
            print(f'{group}: Inserted {len(all_rows)} rows')
            grand_total += len(all_rows)
            if delete_file(file_info['id']):
                print(f'{group}: Deleted from OneDrive')
            else:
                print(f'{group}: Failed to delete from OneDrive')
        else:
            print(f'{group}: No data rows')

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    sync_data = {'last_sync': local_time.isoformat(), 'had_data': grand_total > 0}
    with open(settings.CREDITOR_LAST_SYNC_FILE, 'w') as f:
        json.dump(sync_data, f)

    print(f'Creditor sync complete: {grand_total} total rows')
    return grand_total


def sync_condor_dor_data():
    """Sync Condor+DOR PNL data from OneDrive to PostgreSQL"""
    folder_path = settings.ONEDRIVE_CONDOR_DOR_FOLDER_PATH
    files = list_files_in_folder(folder_path)

    # Filter Excel files
    excel_files = [
        f for f in files
        if f['name'].lower().endswith(('.xlsx', '.xls'))
    ]

    print(f"Found {len(excel_files)} Excel files in Condor+DOR folder")

    # Map filenames to department and branch
    FILE_MAP = {
        'CON': ('CON', 'CON'),
        'DOR BRK': ('BRK', 'DOR'),
        'DOR FEA': ('FEA', 'DOR'),
        'DOR TRX': ('TRX', 'DOR'),
    }

    grand_total = 0

    for file_info in excel_files:
        fname = file_info['name']
        print(f"Processing {fname}...")

        # Determine department and branch from filename
        department = None
        branch = None
        for key, (dept, br) in FILE_MAP.items():
            if key in fname.upper() or key in fname:
                department = dept
                branch = br
                break

        if not department:
            print(f"  Skipping - cannot determine department from filename")
            continue

        print(f"  Department: {department}, Branch: {branch}")

        file_content = download_file(file_info['id'])
        if not file_content:
            print(f"  Failed to download")
            continue

        # Read raw bytes
        if hasattr(file_content, 'read'):
            file_content.seek(0)
            raw = file_content.read()
        else:
            raw = file_content

        # Try openpyxl first, fall back to xlrd
        ws_data = None
        try:
            import io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
            sheet_name = 'GL PL Period Analysis' if 'GL PL Period Analysis' in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            print(f"  Opened with openpyxl - Sheet: {sheet_name}")

            # Find header row with period columns (YYYYMM)
            header_row_idx = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(60, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col).value
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        header_row_idx = row_idx
                        break
                if header_row_idx:
                    break

            if header_row_idx is None:
                print(f"  Error: Could not find header row with period columns")
                continue

            # Get period columns (first occurrence of each)
            month_columns = []
            seen = set()
            for col in range(1, min(60, ws.max_column + 1)):
                val = ws.cell(row=header_row_idx, column=col).value
                if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                    m = int(val)
                    if m not in seen:
                        month_columns.append((col, m))
                        seen.add(m)

            print(f"  Found {len(month_columns)} periods")

            # Parse data rows
            rows = []
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                account_name = ws.cell(row=row_idx, column=3).value
                if not account_name:
                    continue
                account_name = str(account_name).strip()

                for col, period in month_columns:
                    value = ws.cell(row=row_idx, column=col).value
                    if value is not None and isinstance(value, (int, float)):
                        date_str = str(period)
                        date_fixed = yyyymm_to_date(period)
                        rows.append((department, account_name, value, date_str, date_fixed, branch, 'Actual'))
            ws_data = rows

        except Exception as e:
            print(f"  openpyxl failed: {e}, trying xlrd...")
            try:
                wb = xlrd.open_workbook(file_contents=raw)
                sheet_name = 'GL PL Period Analysis' if 'GL PL Period Analysis' in wb.sheet_names() else wb.sheet_names()[0]
                ws = wb.sheet_by_name(sheet_name)
                print(f"  Opened with xlrd - Sheet: {sheet_name}, rows={ws.nrows}")

                # Find header row with period columns
                header_row_idx = None
                for row_idx in range(min(20, ws.nrows)):
                    for col_idx in range(min(60, ws.ncols)):
                        val = ws.cell_value(row_idx, col_idx)
                        if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                            header_row_idx = row_idx
                            break
                    if header_row_idx is not None:
                        break

                if header_row_idx is None:
                    print(f"  Error: Could not find header row with period columns")
                    continue

                # Get period columns
                month_columns = []
                seen = set()
                for col_idx in range(min(60, ws.ncols)):
                    val = ws.cell_value(header_row_idx, col_idx)
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        m = int(val)
                        if m not in seen:
                            month_columns.append((col_idx, m))
                            seen.add(m)

                print(f"  Found {len(month_columns)} periods")

                # Parse data rows
                rows = []
                for row_idx in range(header_row_idx + 1, ws.nrows):
                    account_name = ws.cell_value(row_idx, 2) if ws.ncols > 2 else None
                    if not account_name:
                        continue
                    account_name = str(account_name).strip()

                    for col_idx, period in month_columns:
                        if col_idx >= ws.ncols:
                            continue
                        value = ws.cell_value(row_idx, col_idx)
                        if value is not None and isinstance(value, (int, float)):
                            date_str = str(period)
                            date_fixed = yyyymm_to_date(period)
                            rows.append((department, account_name, value, date_str, date_fixed, branch, 'Actual'))
                ws_data = rows

            except Exception as e2:
                print(f"  xlrd also failed: {e2}")
                continue

        if not ws_data:
            print(f"  No data extracted")
            continue

        # Get unique periods from this file
        periods = set(r[3] for r in ws_data)
        print(f"  Extracted {len(ws_data)} rows across {len(periods)} periods")

        # Delete existing Actual data for this department+branch+periods, then insert
        with connection.cursor() as cur:
            for period in periods:
                cur.execute(
                    "DELETE FROM condor_dor_pnl WHERE department = %s AND branch = %s AND date = %s AND budget_actual = 'Actual'",
                    (department, branch, period)
                )
            print(f"  Deleted old Actual data for {department}/{branch}")

            execute_values(
                cur,
                "INSERT INTO condor_dor_pnl (department, account_name, value, date, date_fixed, branch, budget_actual) VALUES %s",
                ws_data
            )
            print(f"  Inserted {len(ws_data)} rows")

        grand_total += len(ws_data)

    # Save last sync time
    from zoneinfo import ZoneInfo
    local_time = datetime.now(ZoneInfo('Africa/Johannesburg'))
    sync_data = {'last_sync': local_time.isoformat(), 'had_data': grand_total > 0}
    with open(settings.CONDOR_DOR_LAST_SYNC_FILE, 'w') as f:
        json.dump(sync_data, f)

    print(f"\nCondor+DOR sync complete: {grand_total} total rows")
    return grand_total


def get_condor_dor_last_sync():
    """Get the last Condor+DOR sync info"""
    try:
        with open(settings.CONDOR_DOR_LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            if 'last_sync' in data:
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None
