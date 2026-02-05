import os
import io
import json
from datetime import datetime
from openpyxl import load_workbook
from django.conf import settings
from django.db import connection
from psycopg2.extras import execute_values

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
PROGRESS_FILE = '/tmp/sync_progress.json'
# Use settings.LAST_SYNC_FILE for consistency with onedrive_sync


def save_last_sync(rows_count):
    """Save the last sync timestamp"""
    now = datetime.now()
    with open(settings.LAST_SYNC_FILE, 'w') as f:
        json.dump({
            'timestamp': now.isoformat(),
            'time': now.strftime('%H:%M'),
            'date': now.strftime('%B %d, %Y'),
            'rows': rows_count
        }, f)


def get_last_sync():
    """Get the last sync info"""
    try:
        with open(settings.LAST_SYNC_FILE, 'r') as f:
            data = json.load(f)
            # Handle both old format (with 'last_sync' key) and new format (with time/date keys)
            if 'last_sync' in data:
                # OneDrive sync format - convert to display format
                dt = datetime.fromisoformat(data['last_sync'])
                return {
                    'time': dt.strftime('%H:%M'),
                    'date': dt.strftime('%B %d, %Y')
                }
            return data
    except:
        return None

def update_progress(status, message, current=0, total=0):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump({
            'status': status,
            'message': message,
            'current': current,
            'total': total,
            'percent': int((current / total * 100) if total > 0 else 0)
        }, f)

def get_progress():
    try:
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    except:
        return {'status': 'idle', 'message': '', 'current': 0, 'total': 0, 'percent': 0}

def get_drive_service():
    creds = None
    if os.path.exists(settings.GOOGLE_TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(settings.GOOGLE_TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(settings.GOOGLE_CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=8888)
        with open(settings.GOOGLE_TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)

def list_folders(service, parent_id):
    query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def list_xlsx_files(service, folder_id):
    query = f"'{folder_id}' in parents and name contains '.xlsx' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def read_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    file_buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    file_buffer.seek(0)
    return file_buffer

KNOWN_BRANCHES = ['ATL', 'HEC', 'HNL', 'HOU', 'ICS', 'IMP', 'JFK', 'LAX', 'LCL', 'ORD', 'PPG']

def get_branch(filename):
    name = filename.upper().replace('.XLSX', '')
    if 'ALL OVER' in name:
        return None

    # Check if filename starts with a known branch code (e.g., "HNL Analysis - ...")
    for branch in KNOWN_BRANCHES:
        if name.startswith(branch):
            return branch

    # If exact match (e.g., "ATL.xlsx" -> "ATL")
    if name in KNOWN_BRANCHES:
        return name

    return None

def parse_month_code(code):
    try:
        code_str = str(int(code))
        year = int(code_str[:4])
        month = int(code_str[4:])
        if 1 <= month <= 12:
            return datetime(year, month, 1)
    except:
        pass
    return None

def process_xlsx_from_drive(file_buffer, branch):
    data = []
    wb = load_workbook(file_buffer, data_only=True)
    sheet = wb.active

    month_columns = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row and len(row) > 6 and row[3] == 'Debtor':
            for col_idx in range(6, len(row)):
                if row[col_idx]:
                    date_val = parse_month_code(row[col_idx])
                    if date_val:
                        month_columns[col_idx] = date_val
            break

    if not month_columns:
        wb.close()
        return data

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue
        debtor_code = row[3]
        debtor_name = row[4]
        if not debtor_code or debtor_code == 'Debtor':
            continue
        if not isinstance(debtor_code, str):
            continue
        for col_idx, date_val in month_columns.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    value = float(row[col_idx])
                    if value != 0:
                        data.append((
                            debtor_code,
                            debtor_name,
                            value,
                            date_val.strftime('%Y-%m-%d'),
                            date_val.strftime('%Y-%m-%d'),
                            branch
                        ))
                except (ValueError, TypeError):
                    continue

    wb.close()
    return data

def sync_google_drive_data():
    update_progress('running', 'Authenticating with Google Drive...', 0, 100)

    service = get_drive_service()

    update_progress('running', 'Scanning Google Drive folder...', 5, 100)
    year_folders = list_folders(service, settings.GOOGLE_DRIVE_FOLDER_ID)

    # Count total files
    files_to_process = []
    for year_folder in sorted(year_folders, key=lambda x: x['name']):
        xlsx_files = list_xlsx_files(service, year_folder['id'])
        xlsx_files = [f for f in xlsx_files if 'ALL OVER' not in f['name'].upper()]
        for f in xlsx_files:
            branch = get_branch(f['name'])
            if branch:
                files_to_process.append({
                    'year': year_folder['name'],
                    'file': f,
                    'branch': branch
                })

    update_progress('running', f'Found {len(files_to_process)} files to process...', 10, 100)

    all_data = []
    processed = 0

    for item in files_to_process:
        filename = item['file']['name']
        file_id = item['file']['id']
        branch = item['branch']
        year = item['year']

        try:
            update_progress('running', f"Reading {year}/{filename}...", 10 + int(processed / len(files_to_process) * 60), 100)
            file_buffer = read_file(service, file_id)

            update_progress('running', f"Processing {year}/{filename}...", 10 + int(processed / len(files_to_process) * 60), 100)
            data = process_xlsx_from_drive(file_buffer, branch)
            all_data.extend(data)
            processed += 1
        except Exception as e:
            print(f"Error: {filename}: {e}")
            processed += 1

    update_progress('running', f'Collected {len(all_data)} rows. Creating table...', 75, 100)

    with connection.cursor() as cursor:
        # Check if table exists and has the unique constraint
        cursor.execute('''
            SELECT COUNT(*) FROM information_schema.table_constraints
            WHERE table_name = 'turnover_data'
            AND constraint_type = 'UNIQUE'
        ''')
        has_constraint = cursor.fetchone()[0] > 0

        if not has_constraint:
            # Drop old table and recreate with unique constraint (one-time migration)
            update_progress('running', 'Migrating table structure...', 76, 100)
            cursor.execute("DROP TABLE IF EXISTS turnover_data CASCADE")

        # Create table if not exists with unique constraint for upsert
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS turnover_data (
                id SERIAL PRIMARY KEY,
                debtor VARCHAR(100),
                debtor_name VARCHAR(255),
                value DECIMAL(15,2),
                date DATE,
                date_fixed VARCHAR(20),
                branch VARCHAR(50),
                UNIQUE(debtor, date, branch)
            )
        ''')

        # Create index for faster lookups if not exists
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_turnover_debtor_date_branch
            ON turnover_data(debtor, date, branch)
        ''')

        update_progress('running', f'Upserting {len(all_data)} rows...', 80, 100)

        # Bulk upsert using execute_values - updates existing records, inserts new ones
        batch_size = 5000
        total = len(all_data)

        for i in range(0, total, batch_size):
            batch = all_data[i:i+batch_size]
            execute_values(
                cursor,
                '''INSERT INTO turnover_data (debtor, debtor_name, value, date, date_fixed, branch)
                   VALUES %s
                   ON CONFLICT (debtor, date, branch)
                   DO UPDATE SET
                       debtor_name = EXCLUDED.debtor_name,
                       value = EXCLUDED.value,
                       date_fixed = EXCLUDED.date_fixed''',
                batch,
                page_size=batch_size
            )
            percent = 80 + int((i + batch_size) / total * 19)
            update_progress('running', f'Processed {min(i+batch_size, total)}/{total} rows...', min(percent, 99), 100)

    save_last_sync(len(all_data))
    update_progress('complete', f'Done! {len(all_data)} rows imported.', 100, 100)
    return {'rows': len(all_data)}
