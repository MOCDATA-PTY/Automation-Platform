"""Inspect Book1 1.xlsx"""
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\berna\OneDrive\Desktop\Automation-Platform-master\Automation-Platform-master\automations\templates\Book1 1.xlsx', read_only=True)
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print()
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
    print(f'Row {i}: {list(row)}')
wb.close()
