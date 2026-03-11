"""Import Book1 3.xlsx into useu_contacts (append, no delete)."""
import django, os
os.environ['DJANGO_SETTINGS_MODULE'] = 'automations.settings'
django.setup()

import openpyxl
from dashboard.models import USEUContact

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Book1 3.xlsx')

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
ws = wb.active

def safe(val):
    if val is None:
        return ''
    s = str(val).strip().lstrip('\t')
    return s

# Print headers
headers = [str(c) if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
print("Columns:", len(headers))
for i, h in enumerate(headers):
    print(f"  {i}: {h}")

batch = []
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    email_val = safe(row[13] if len(row) > 13 else '')
    status = safe(row[26] if len(row) > 26 else 'Active')
    if '<' in email_val and '>' in email_val:
        status = 'Faulty Data'
    if not status:
        status = 'Active'

    batch.append(USEUContact(
        title=safe(row[0] if len(row) > 0 else ''),
        org_name=safe(row[1] if len(row) > 1 else ''),
        default=safe(row[2] if len(row) > 2 else ''),
        contact_name=safe(row[3] if len(row) > 3 else ''),
        mode=safe(row[4] if len(row) > 4 else ''),
        attach=safe(row[5] if len(row) > 5 else ''),
        job_title=safe(row[6] if len(row) > 6 else ''),
        address=safe(row[7] if len(row) > 7 else ''),
        branch=safe(row[8] if len(row) > 8 else ''),
        unloco=safe(row[9] if len(row) > 9 else ''),
        city=safe(row[10] if len(row) > 10 else ''),
        phone=safe(row[11] if len(row) > 11 else ''),
        fax=safe(row[12] if len(row) > 12 else ''),
        email=email_val,
        sales_rep=safe(row[14] if len(row) > 14 else ''),
        touchpoint_1=safe(row[15] if len(row) > 15 else ''),
        touchpoint_2=safe(row[16] if len(row) > 16 else ''),
        touchpoint_3=safe(row[17] if len(row) > 17 else ''),
        touchpoint_4=safe(row[18] if len(row) > 18 else ''),
        touchpoint_5=safe(row[19] if len(row) > 19 else ''),
        touchpoint_6=safe(row[20] if len(row) > 20 else ''),
        touchpoint_7=safe(row[21] if len(row) > 21 else ''),
        touchpoint_8=safe(row[22] if len(row) > 22 else ''),
        touchpoint_9=safe(row[23] if len(row) > 23 else ''),
        touchpoint_10=safe(row[24] if len(row) > 24 else ''),
        last_touch=safe(row[25] if len(row) > 25 else ''),
        journey_status=status,
        tp1_sent_on=safe(row[27] if len(row) > 27 else ''),
        tp2_sent_on=safe(row[28] if len(row) > 28 else ''),
        tp3_sent_on=safe(row[29] if len(row) > 29 else ''),
        tp5_sent_on=safe(row[30] if len(row) > 30 else ''),
        tp4_sent_on=safe(row[31] if len(row) > 31 else ''),
        tp6_sent_on=safe(row[32] if len(row) > 32 else ''),
        tp7_sent_on=safe(row[33] if len(row) > 33 else ''),
        tp8_sent_on=safe(row[34] if len(row) > 34 else ''),
        tp9_sent_on=safe(row[35] if len(row) > 35 else ''),
        tp10_sent_on=safe(row[36] if len(row) > 36 else ''),
        status=status,
    ))
    count += 1
    if count % 5000 == 0:
        print(f"  Read {count} rows...")

wb.close()

print(f"\nBefore import: {USEUContact.objects.count()} records in DB")
print(f"Appending {count} records...")
USEUContact.objects.bulk_create(batch, batch_size=500)
print(f"Done! Total in DB: {USEUContact.objects.count()}")
