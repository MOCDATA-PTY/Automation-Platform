"""Local DFW sync - downloads from OneDrive, processes, loads into DB.
Does NOT delete the file from OneDrive.
"""
import json, msal, requests, io, re, os, sys
from datetime import datetime
from collections import defaultdict
import xlrd
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# --- Config ---
CLIENT_ID = '43fbe5a9-6b5b-4c81-9067-7aff9ac3ed5a'
TENANT_ID = 'b1504b1d-d096-409a-a0f0-6cc546dde993'
CLIENT_SECRET = 'w6B8Q~W3ac9klXa8NkMDo4cPNyOsjEryVL5TwdhQ'
AUTHORITY = f'https://login.microsoftonline.com/{TENANT_ID}'
SCOPES = ['Files.Read.All', 'Files.ReadWrite.All']
GRAPH = 'https://graph.microsoft.com/v1.0'
DFW_FOLDER = '/Automation Platform/DFW Financial Analysis Report'

DB_CONFIG = {
    'dbname': 'turnover_data',
    'user': 'powerbi',
    'password': 'your_secure_password',
    'host': '167.88.43.168',
    'port': '5432'
}


def yyyymm_to_date(month_yyyymm):
    year = month_yyyymm // 100
    month = month_yyyymm % 100
    return f"{year}-{month:02d}-01"


def calculate_week_number(report_date, month_yyyymm):
    data_year = month_yyyymm // 100
    data_month = month_yyyymm % 100
    report_year = report_date.year
    report_month = report_date.month

    if (data_year < report_year) or (data_year == report_year and data_month < report_month):
        if data_month == 2:
            return 4
        return 5

    day = report_date.day
    if day <= 7:
        return None

    if data_month == 2:
        if day <= 14: return 1
        elif day <= 21: return 2
        elif day <= 27: return 3
        else: return 4

    if day <= 14: return 1
    elif day <= 21: return 2
    elif day <= 28: return 3
    else: return 4


def process_ppg_excel_file(file_content, filename):
    """Exact copy of production process_ppg_excel_file"""
    rows = []
    try:
        date_match = re.search(r'(\d{2})\s+(\w+)\s+(\d{4})', filename)
        if date_match:
            day = int(date_match.group(1))
            month_name = date_match.group(2)
            year = int(date_match.group(3))
            report_date = datetime.strptime(f"{day} {month_name} {year}", "%d %B %Y")
            print(f"  Report date: {report_date.strftime('%Y-%m-%d')}")
        else:
            print(f"  Warning: Could not extract report date from filename, using today")
            report_date = datetime.now()

        file_content.seek(0)

        try:
            wb = openpyxl.load_workbook(file_content, data_only=True)
            ws = wb['GL PL Period Analysis']
            print(f"  Sheet: GL PL Period Analysis (openpyxl)")
            print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")

            header_row_idx = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(30, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col).value
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        header_row_idx = row_idx
                        break
                if header_row_idx:
                    break

            if header_row_idx is None:
                print(f"  Error: Could not find header row with month columns")
                return []

            print(f"  Found header row at row {header_row_idx}")

            month_columns = []
            seen_months = set()
            for col_num in range(1, min(60, ws.max_column + 1)):
                val = ws.cell(row=header_row_idx, column=col_num).value
                if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                    month_val = int(val)
                    if month_val not in seen_months:
                        month_columns.append((col_num, month_val))
                        seen_months.add(month_val)

            print(f"  Found {len(month_columns)} month columns: {month_columns}")

            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                if not any(ws[row_idx]):
                    continue

                account_number = str(ws.cell(row=row_idx, column=2).value).strip() if ws.cell(row=row_idx, column=2).value else None
                account_name = str(ws.cell(row=row_idx, column=3).value).strip() if ws.cell(row=row_idx, column=3).value else None

                if not account_number and not account_name:
                    continue

                for col_num, month_yyyymm in month_columns:
                    value = ws.cell(row=row_idx, column=col_num).value
                    if value is not None and isinstance(value, (int, float)):
                        week = calculate_week_number(report_date, month_yyyymm)
                        if week is None:
                            continue
                        week_label = f"Week {week}"
                        date_fixed = yyyymm_to_date(month_yyyymm)
                        budget_actual = "Actual"
                        rows.append(("PPG", account_name, value, month_yyyymm, date_fixed, budget_actual, week_label, report_date.strftime('%Y-%m-%d')))

        except Exception as xlsx_error:
            print(f"  openpyxl failed ({xlsx_error}), trying xlrd...")
            file_content.seek(0)
            wb = xlrd.open_workbook(file_contents=file_content.read())
            ws = wb.sheet_by_name('GL PL Period Analysis')
            print(f"  Sheet: GL PL Period Analysis (xlrd)")
            print(f"  Rows: {ws.nrows}, Columns: {ws.ncols}")

            header_row_idx = None
            for row_idx in range(min(20, ws.nrows)):
                for col_idx in range(min(30, ws.ncols)):
                    val = ws.cell_value(row_idx, col_idx)
                    if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                        header_row_idx = row_idx
                        break
                if header_row_idx is not None:
                    break

            if header_row_idx is None:
                print(f"  Error: Could not find header row with month columns")
                return []

            print(f"  Found header row at row {header_row_idx + 1}")

            month_columns = []
            seen_months = set()
            for col_idx in range(min(60, ws.ncols)):
                val = ws.cell_value(header_row_idx, col_idx)
                if isinstance(val, (int, float)) and 202000 <= val <= 209999:
                    month_val = int(val)
                    if month_val not in seen_months:
                        month_columns.append((col_idx, month_val))
                        seen_months.add(month_val)

            print(f"  Found {len(month_columns)} month columns: {month_columns}")

            for row_idx in range(header_row_idx + 1, ws.nrows):
                row = ws.row_values(row_idx)
                if not any(row):
                    continue

                account_number = str(ws.cell_value(row_idx, 1)).strip() if ws.ncols > 1 and ws.cell_value(row_idx, 1) else None
                account_name = str(ws.cell_value(row_idx, 2)).strip() if ws.ncols > 2 and ws.cell_value(row_idx, 2) else None

                if not account_number and not account_name:
                    continue

                for col_idx, month_yyyymm in month_columns:
                    if col_idx >= ws.ncols:
                        continue
                    value = ws.cell_value(row_idx, col_idx)
                    if value is not None and isinstance(value, (int, float)):
                        week = calculate_week_number(report_date, month_yyyymm)
                        if week is None:
                            continue
                        week_label = f"Week {week}"
                        date_fixed = yyyymm_to_date(month_yyyymm)
                        budget_actual = "Actual"
                        rows.append(("PPG", account_name, value, month_yyyymm, date_fixed, budget_actual, week_label, report_date.strftime('%Y-%m-%d')))

    except Exception as e:
        print(f"  Error processing file {filename}: {e}")
        import traceback
        traceback.print_exc()
        return []

    return rows


def main():
    # --- Step 1: Refresh token ---
    print("=" * 60)
    print("STEP 1: Refreshing OneDrive token...")

    token_file = 'onedrive_token.json'
    if os.path.exists(token_file):
        with open(token_file) as f:
            token_data = json.load(f)
        refresh_token = token_data['refresh_token']
    else:
        # Use the refresh token from the server paste
        refresh_token = '1.AU4AHUtQsZbQmkCg8GzFRt3pk6nl-0Nba4FMkGd6_5rD7VoBAIhOAA.BQABAwEAAAADAOz_BQD0_0V2b1N0c0FydGlmYWN0cwIAAAAAAD0qOO_S50QqsUZ_tINzZtyAwAgaWWc2Jvt05VMgyT1t4TavhFixmnQ2QTCmOYfUAYsyzYnhGoKlJB2urOTi2rIazCQyIq928Rxu8mZ3mOei0ByRf__BjzO_7wyT8zaxpN4UHQvbrObcA3sB_7r3eTnsVzHJkBHLuMHeJIo2tvgq3Emgz9BMY86ORQ2v2Nk0IsTs2KFKhY1ELah4fEy0Q-9lI8zafJRcSI0NQHXEavwoJeOSbqV5Qbu-F2B25s5vymQsfc6q0XfpbTaUDDUgSlL7eyGk7CZQXMPxIO2jAzlO6cOUWMZuBTm9UrGAZd7BHGU-fygSPgFQaTkWyQ8ROV5b_YtNxy2pjGr_eW-xaK6xot4eiFMrrrgXfdx0sjMq4Vlaa3p9hDp3BlsWyS7Ex6Zx8v0Ug9o8o2JpSFERYCwkC20eh0e-184FYDYIzy5l9mDBAMIlfjUcXkO33FOc9s7nFLDOOXtYCYM3-aCIlsy3Pu4Z5JMWxcBfUBbdnckWKuikugv6rDr9ijd1Pu4fV5MnYow1XXtLqNnzavzBXKgHEx20KrjZYDtPr5RlizJm7b1elcVCzm8S3eKFolTt8W5yfx5cMAcJdMdeBQKczjN2YwY57atm6BC9XhAy5l-YXhai1AZE2mcDpuEwvo1Zr2-Xn6Ro8y1Q9b87JrlF80gfc8Kb_SLl2KTfNlfPVg2ZwAnzBNBfY_Mz4tvQojNV23qggX-Ltd7oXy3dC8_F0yzyZkoi-QQrOQGCP-hKbiIdVH4W8hPsjz6_cNS4ulgMfETc-YM6SjZ-p1MxeXlCkaJNdNvPyrLM4Iuu51ZvDTMS5EOZYWqQe0s9BNY8Qj4YdfmPaWzvgPExiI-Ag_smCANrFB0yvLPsIYWWgxapOE7awBJgP1i8TorRyA9vr8StGdN6nbFMeuQkksOiNUpVGBsg_BMdboxMrcbUfhpuUYAmyGDzyfh54Z81yMulp2L9G3FLTwb_dl0YooK5gRizVrCK8TBtBEI0lq1eu7qWC2VFQSaPvmwKq-0lkhxM6--OMT69AB6BbufaZMr7h5xmVIwY_AxE7M9E5r0OsTbMbhMk9GAOVC0M9NVXMCkXo-r3bhzmNyBVmRORBKGWAST4018cn1TiZd72URR57CoOukAFpLhEOJMB0hsMjOrpp01SISE9rTa_bfc4xQDPMQCLQC-Lf7Wx_vx7QaSEcNGbU7i8BegB8u2NXi_4_hdF7HkHolEl-JmYYqn7HjwtjbuC8AGpg5IO_fSYh51Q9gdS7tmWVAHe2gZ3lwUUDbGYOVkmFtYdNMM7B18nzhzaZnUiYiglfzKA1Tdqh9FdZCMQL4qdiRkHLmnkzBZ9xCVT6asyV1PTJpUm_S73ZGmy1Hcw0qUVbqp2zqjb8x7VyEieXBaSsrmRBNfHDcayb0tjBA'
        token_data = {}

    app = msal.ConfidentialClientApplication(CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)
    result = app.acquire_token_by_refresh_token(refresh_token, scopes=SCOPES)

    if 'access_token' not in result:
        print(f"FAILED: {result.get('error')}: {result.get('error_description')}")
        sys.exit(1)

    access_token = result['access_token']
    print("Token refreshed OK")

    # Save updated token locally for future use
    token_data.update(result)
    with open(token_file + '.tmp', 'w') as f:
        json.dump(token_data, f, indent=2)
    os.replace(token_file + '.tmp', token_file)
    print("Token saved to onedrive_token.json")

    headers = {'Authorization': f'Bearer {access_token}'}

    # --- Step 2: List & download DFW files ---
    print("\n" + "=" * 60)
    print("STEP 2: Listing DFW folder...")

    folder_encoded = DFW_FOLDER.strip('/')
    url = f'{GRAPH}/me/drive/root:/{folder_encoded}:/children'
    resp = requests.get(url, headers=headers, timeout=30)

    if resp.status_code != 200:
        print(f"Error listing folder: {resp.status_code} {resp.text}")
        sys.exit(1)

    items = resp.json().get('value', [])
    excel_files = [item for item in items if item['name'].lower().endswith(('.xlsx', '.xls'))]
    print(f"Found {len(excel_files)} Excel files")

    if not excel_files:
        print("No Excel files to process")
        sys.exit(0)

    for item in excel_files:
        print(f"  {item['name']} ({item.get('size', 0) / 1024:.1f} KB)")

    # --- Download and process ---
    all_rows = []

    for item in excel_files:
        fname = item['name']
        fid = item['id']
        print(f"\nDownloading: {fname}...")

        dl_url = f'{GRAPH}/me/drive/items/{fid}/content'
        resp = requests.get(dl_url, headers=headers, timeout=120)
        if resp.status_code != 200:
            print(f"  Download failed: HTTP {resp.status_code}")
            continue

        file_content = io.BytesIO(resp.content)
        print(f"  Downloaded {len(resp.content)} bytes")

        # Process like PPG then replace division with DFW
        rows = process_ppg_excel_file(file_content, fname)
        rows = [("DFW",) + row[1:] for row in rows]
        all_rows.extend(rows)
        print(f"  Extracted {len(rows)} DFW records")

    print(f"\n{'=' * 60}")
    print(f"STEP 3: Loading {len(all_rows)} records into dfw_pnl...")

    if not all_rows:
        print("No records to insert")
        sys.exit(0)

    # Show sample data
    print(f"\nSample rows (first 5):")
    for r in all_rows[:5]:
        print(f"  {r}")

    # --- Step 3: Insert into database ---
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    cur = conn.cursor()

    unique_month_weeks = set((row[3], row[6]) for row in all_rows)
    unique_months = set(row[3] for row in all_rows)

    print(f"\nDEBUG - Before deletion:")
    for month in sorted(unique_months):
        cur.execute("""
            SELECT week, COUNT(*)
            FROM dfw_pnl
            WHERE date = %s AND budget_actual = 'Actual'
            GROUP BY week ORDER BY week
        """, (month,))
        weeks = cur.fetchall()
        if weeks:
            week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
            print(f"  Month {month}: {week_summary}")
        else:
            print(f"  Month {month}: (no existing data)")

    print(f"\nWill DELETE and replace these (month, week) combinations:")
    for month, week in sorted(unique_month_weeks):
        print(f"  - {month}, {week}")

    # Delete old Actual week records
    for month, week in unique_month_weeks:
        cur.execute("DELETE FROM dfw_pnl WHERE date = %s AND week = %s AND budget_actual = 'Actual'", (month, week))
    print(f"\nDeleted old Actual week records for {len(unique_month_weeks)} (month, week) combos")

    # Delete old Actual Total records
    for month in unique_months:
        cur.execute("DELETE FROM dfw_pnl WHERE date = %s AND week = 'Total' AND budget_actual = 'Actual'", (month,))
    print(f"Deleted old Actual 'Total' records for {len(unique_months)} months")

    # Insert new week data
    week_query = "INSERT INTO dfw_pnl (division, account_name, value, date, date_fixed, budget_actual, week, report_date) VALUES %s"
    execute_values(cur, week_query, all_rows)
    print(f"Inserted {len(all_rows)} week records")

    # Duplicate as Total rows per month
    rows_by_month = defaultdict(list)
    for row in all_rows:
        rows_by_month[row[3]].append(row)

    total_count = 0
    for month, month_rows in rows_by_month.items():
        total_rows = [(row[0], row[1], row[2], row[3], row[4], row[5], 'Total', row[7]) for row in month_rows]
        execute_values(cur, week_query, total_rows)
        total_count += len(total_rows)
        print(f"  Inserted {len(total_rows)} Total records for month {month}")

    print(f"Total: {total_count} Total records inserted")

    # DEBUG: Show weeks after insertion
    print(f"\nDEBUG - After insertion:")
    for month in sorted(unique_months):
        cur.execute("""
            SELECT week, COUNT(*)
            FROM dfw_pnl
            WHERE date = %s AND budget_actual = 'Actual'
            GROUP BY week ORDER BY week
        """, (month,))
        weeks = cur.fetchall()
        if weeks:
            week_summary = ", ".join([f"{w}({c})" for w, c in weeks])
            print(f"  Month {month}: {week_summary}")

    cur.close()
    conn.close()

    print(f"\n{'=' * 60}")
    print(f"DONE: Imported {len(all_rows)} week records + {total_count} Total records into dfw_pnl")
    print(f"NOTE: File NOT deleted from OneDrive (kept as requested)")


if __name__ == '__main__':
    main()
