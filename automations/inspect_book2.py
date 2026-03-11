"""Inspect Book1 2.xlsx to see column structure and sample data."""
import openpyxl

wb = openpyxl.load_workbook('Book1 2.xlsx', read_only=True)
ws = wb.active

# Print headers
headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
    headers.append(str(cell) if cell else '')

print("=== COLUMNS ===")
for i, h in enumerate(headers):
    print(f"  Col {i}: {h}")
print(f"\nTotal columns: {len(headers)}")

# Print first 5 data rows
for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
    vals = [str(v)[:50] if v else '' for v in row]
    print(f"\nRow {ri+2}:")
    for i, v in enumerate(vals):
        if v:
            hdr = headers[i] if i < len(headers) else '?'
            print(f"  [{i}] {hdr}: {v}")

# Count total rows
total = 0
for _ in ws.iter_rows(min_row=2, values_only=True):
    total += 1
print(f"\nTotal data rows: {total}")

wb.close()
