"""
Standalone turnover sync script - pulls all Excel files from OneDrive and upserts to PostgreSQL.
Processes all .xlsx/.xls files in the root folder, skips year subfolders.
Deletes processed files after successful sync.

Usage: python3 sync_turnover.py
"""
import json
import os
import io
import re
import msal
import requests
import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from datetime import datetime

# --- OneDrive config ---
CLIENT_ID = '43fbe5a9-6b5b-4c81-9067-7aff9ac3ed5a'
TENANT_ID = 'b1504b1d-d096-409a-a0f0-6cc546dde993'
CLIENT_SECRET = 'w6B8Q~W3ac9klXa8NkMDo4cPNyOsjEryVL5TwdhQ'
SCOPES = ['Files.Read.All', 'Files.ReadWrite.All']
TOKEN_FILE = os.path.join(os.path.dirname(__file__), 'onedrive_token.json')
GRAPH_API_ENDPOINT = 'https://graph.microsoft.com/v1.0'
FOLDER_PATH = '/Automation Platform/Turn Over Automation Report'

# --- DB config ---
DB = {
    'host': '167.88.43.168',
    'port': 5432,
    'dbname': 'turnover_data',
    'user': 'powerbi',
    'password': 'your_secure_password',
}

# --- Log file ---
LOG_FILE = os.path.join(os.path.dirname(__file__), 'sync_log.txt')


def log(msg):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, 'a') as f:
        f.write(line + '\n')


def get_access_token():
    if not os.path.exists(TOKEN_FILE):
        log("ERROR: No token file found. Run auth.py first.")
        return None
    with open(TOKEN_FILE, 'r') as f:
        token_data = json.load(f)
    refresh_token = token_data.get('refresh_token')
    if refresh_token:
        app = msal.ConfidentialClientApplication(
            CLIENT_ID,
            authority=f"https://login.microsoftonline.com/{TENANT_ID}",
            client_credential=CLIENT_SECRET
        )
        result = app.acquire_token_by_refresh_token(refresh_token, scopes=SCOPES)
        if 'access_token' in result:
            with open(TOKEN_FILE, 'w') as f:
                json.dump(result, f, indent=2)
            return result['access_token']
        else:
            log(f"Token refresh failed: {result.get('error_description', 'unknown')}")
    return token_data.get('access_token')


def get_branch_from_fw(filename):
    m = re.search(r'-\s*([A-Z\-]+?)(\d{4}-\d{2}-\d{2})', filename)
    return m.group(1) if m else None


def extract_report_date(ws):
    try:
        row11_text = str(ws.cell(row=11, column=2).value)
        date_match = re.search(r'(\d{2}-[A-Za-z]{3}-\d{2})\s+(\d{1,2}:\d{2}\s+[AP]M)', row11_text)
        if date_match:
            report_date = datetime.strptime(
                f"{date_match.group(1)} {date_match.group(2)}", "%d-%b-%y %I:%M %p"
            )
            return report_date.strftime('%Y-%m-%d')
    except:
        pass
    return None


def process_file(ws, branch, report_date):
    rows = []
    # Find month columns in row 13
    month_columns = []
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=13, column=col_idx).value
        if isinstance(v, (int, float)) and 202000 <= int(v) <= 209912:
            month_columns.append((col_idx, int(v)))

    if not month_columns:
        # LONG format: C1=Debtor, C2=Name, C3=month, C4=value, data from row 2
        for row_idx in range(2, ws.max_row + 1):
            debtor = ws.cell(row=row_idx, column=1).value
            if not debtor:
                continue
            name = ws.cell(row=row_idx, column=2).value
            date_val = ws.cell(row=row_idx, column=3).value
            value = ws.cell(row=row_idx, column=4).value
            if not date_val or not value:
                continue
            value = float(value)
            if value == 0:
                continue
            try:
                date_int = int(float(date_val)) if not isinstance(date_val, (int, float)) else int(date_val)
            except:
                continue
            year, month = date_int // 100, date_int % 100
            date_str = f"{year:04d}-{month:02d}-01"
            rows.append((str(debtor).strip(), str(name).strip() if name else '',
                         date_str, value, date_str, branch, report_date))
        return rows

    # WIDE format: row 13 headers with month columns, debtor/name columns, data from row 14
    debtor_col, name_col = 4, 5
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=13, column=col_idx).value
        if v and str(v).strip().lower() == 'debtor':
            debtor_col = col_idx
        if v and str(v).strip().lower() in ('debtor name', 'name'):
            name_col = col_idx

    for row_idx in range(14, ws.max_row + 1):
        debtor = ws.cell(row=row_idx, column=debtor_col).value
        if not debtor:
            continue
        debtor = str(debtor).strip()
        name = ws.cell(row=row_idx, column=name_col).value
        name = str(name).strip() if name else ''

        for col_idx, month_val in month_columns:
            value = ws.cell(row=row_idx, column=col_idx).value
            if value and isinstance(value, (int, float)) and value != 0:
                year, month = month_val // 100, month_val % 100
                date_str = f"{year:04d}-{month:02d}-01"
                rows.append((debtor, name, date_str, float(value), date_str, branch, report_date))
    return rows


def delete_file(token, file_id):
    headers = {'Authorization': f'Bearer {token}'}
    resp = requests.delete(f"{GRAPH_API_ENDPOINT}/me/drive/items/{file_id}", headers=headers)
    return resp.status_code == 204


def main():
    log("=== Sync started ===")

    # Get token
    token = get_access_token()
    if not token:
        log("ERROR: Could not get access token")
        return
    headers = {'Authorization': f'Bearer {token}'}

    # List files
    url = f'{GRAPH_API_ENDPOINT}/me/drive/root:{FOLDER_PATH}:/children'
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        log(f"ERROR: Failed to list folder: {response.status_code}")
        return
    all_items = response.json().get('value', [])
    # Get all Excel files in the root folder, skip historical year archives (20XX_*.xlsx)
    excel_files = [
        f for f in all_items
        if 'file' in f
        and f['name'].lower().endswith(('.xlsx', '.xls'))
        and not re.match(r'^20\d{2}_', f['name'])  # Skip 2020_ATL.xlsx, 2021_HNL.xlsx, etc.
    ]

    if not excel_files:
        log("No Excel files found. Nothing to do.")
        return

    log(f"Found {len(excel_files)} Excel files to process")

    # Process each file
    all_rows = []
    processed_ids = []
    for file_info in excel_files:
        branch = get_branch_from_fw(file_info['name'])
        if not branch:
            log(f"  SKIP (no branch): {file_info['name']}")
            continue

        try:
            dl_resp = requests.get(
                f"{GRAPH_API_ENDPOINT}/me/drive/items/{file_info['id']}/content",
                headers=headers
            )
            content = io.BytesIO(dl_resp.content)
            wb = openpyxl.load_workbook(content, data_only=True)
            ws = wb.active
            report_date = extract_report_date(ws)
            rows = process_file(ws, branch, report_date)
            all_rows.extend(rows)
            processed_ids.append(file_info['id'])
            log(f"  {branch}: {len(rows)} records (report_date={report_date})")
        except Exception as e:
            log(f"  ERROR processing {branch}: {e}")

    if not all_rows:
        log("No records extracted.")
        return

    log(f"Total records to upsert: {len(all_rows)}")

    # Upsert to DB
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM turnover_data")
    before = cur.fetchone()[0]

    query = """
        INSERT INTO turnover_data (debtor, debtor_name, date, value, date_fixed, branch, report_date)
        VALUES %s
        ON CONFLICT (debtor, date, branch)
        DO UPDATE SET
            value = EXCLUDED.value,
            debtor_name = EXCLUDED.debtor_name,
            date_fixed = EXCLUDED.date_fixed,
            report_date = EXCLUDED.report_date
        WHERE
            turnover_data.report_date IS NULL
            OR EXCLUDED.report_date IS NULL
            OR EXCLUDED.report_date >= turnover_data.report_date
    """
    execute_values(cur, query, all_rows)
    conn.commit()

    cur.execute("SELECT COUNT(*) FROM turnover_data")
    after = cur.fetchone()[0]
    conn.close()

    log(f"DB: {before} -> {after} rows (+{after - before} new)")

    # Delete processed Excel files from OneDrive
    deleted = 0
    for fid in processed_ids:
        if delete_file(token, fid):
            deleted += 1
    log(f"Deleted {deleted}/{len(processed_ids)} Excel files from OneDrive")
    log("=== Sync complete ===\n")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log(f"FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
