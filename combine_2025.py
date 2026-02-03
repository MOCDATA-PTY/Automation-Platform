from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import sqlite3

# Input and output paths
input_folder = "/home/ethan/Desktop/Gork Test/2025"
output_excel = "/home/ethan/Desktop/Gork Test/2025/2025data.xlsx"
output_db = "/home/ethan/Desktop/Gork Test/2025/2025data.db"

# Month mapping for wide format (column index to date)
month_map_wide = {
    6: datetime(2025, 1, 1),
    7: datetime(2025, 2, 1),
    8: datetime(2025, 3, 1),
    9: datetime(2025, 4, 1),
    10: datetime(2025, 5, 1),
    11: datetime(2025, 6, 1),
    12: datetime(2025, 7, 1),
    13: datetime(2025, 8, 1),
    14: datetime(2025, 9, 1),
    15: datetime(2025, 10, 1),
    16: datetime(2025, 11, 1),
    17: datetime(2025, 12, 1),
}

# Month mapping for long format (attribute code to date)
month_map_long = {
    '202501': datetime(2025, 1, 1),
    '202502': datetime(2025, 2, 1),
    '202503': datetime(2025, 3, 1),
    '202504': datetime(2025, 4, 1),
    '202505': datetime(2025, 5, 1),
    '202506': datetime(2025, 6, 1),
    '202507': datetime(2025, 7, 1),
    '202508': datetime(2025, 8, 1),
    '202509': datetime(2025, 9, 1),
    '202510': datetime(2025, 10, 1),
    '202511': datetime(2025, 11, 1),
    '202512': datetime(2025, 12, 1),
}

def get_branch(filename):
    name = filename.upper().replace('.XLSX', '')
    if 'ALL OVER' in name:
        return None
    return name

def detect_format(sheet):
    """Detect if file is wide format or long format"""
    for row in sheet.iter_rows(min_row=1, max_row=2, values_only=True):
        if row and len(row) >= 4:
            if row[0] == 'Debtor' and row[2] == 'Attribute' and row[3] == 'Value':
                return 'long'
            if row[2] and str(row[2]).startswith('2025'):
                return 'long'
    return 'wide'

def process_wide_format(sheet, branch):
    """Process wide format files (ATL, HOU, etc.)"""
    data = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue

        debtor_code = row[3]
        debtor_name = row[4]

        if not debtor_code or debtor_code == 'Debtor':
            continue
        if not isinstance(debtor_code, str):
            continue

        for col_idx, date_val in month_map_wide.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    value = float(row[col_idx])
                    if value != 0:
                        data.append({
                            'Debtor': debtor_code,
                            'Debtor Name': debtor_name,
                            'Value': value,
                            'Date': date_val,
                            'Date Fixed': date_val.strftime('%Y-%m-%d'),
                            'Branch': branch,
                            'Date_PowerBI': date_val
                        })
                except (ValueError, TypeError):
                    continue
    return data

def process_long_format(sheet, branch):
    """Process long format files (HEC, ICS)"""
    data = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 4:
            continue

        debtor_code = row[0]
        debtor_name = row[1]
        attribute = str(row[2]) if row[2] else ''
        value = row[3]

        if not debtor_code or debtor_code == 'Debtor':
            continue
        if not isinstance(debtor_code, str):
            continue
        if attribute not in month_map_long:
            continue

        try:
            value = float(value)
            if value != 0:
                date_val = month_map_long[attribute]
                data.append({
                    'Debtor': debtor_code,
                    'Debtor Name': debtor_name,
                    'Value': value,
                    'Date': date_val,
                    'Date Fixed': date_val.strftime('%Y-%m-%d'),
                    'Branch': branch,
                    'Date_PowerBI': date_val
                })
        except (ValueError, TypeError):
            continue
    return data

all_cleaned_data = []

# Get all xlsx files
files = [f for f in os.listdir(input_folder) if f.lower().endswith('.xlsx')]
files = [f for f in files if 'Cleaned' not in f and '2025data' not in f]

print(f"Found {len(files)} Excel files to process")

for filename in files:
    branch = get_branch(filename)
    if branch is None:
        print(f"Skipping: {filename}")
        continue

    filepath = os.path.join(input_folder, filename)

    try:
        wb = load_workbook(filepath, data_only=True)
        sheet = wb.active

        file_format = detect_format(sheet)
        print(f"Processing: {filename} (Branch: {branch}, Format: {file_format})")

        if file_format == 'long':
            data = process_long_format(sheet, branch)
        else:
            data = process_wide_format(sheet, branch)

        all_cleaned_data.extend(data)
        wb.close()
    except Exception as e:
        print(f"Error processing {filename}: {e}")

# Sort by Debtor, then by Date
all_cleaned_data.sort(key=lambda x: (x['Debtor'], x['Date']))

# === Save to Excel ===
out_wb = Workbook()
out_sheet = out_wb.active
out_sheet.title = "2025data"

headers = ['Debtor', 'Debtor Name', 'Value', 'Date', 'Date Fixed', 'Branch', 'Date_PowerBI']
out_sheet.append(headers)

for row_data in all_cleaned_data:
    out_sheet.append([row_data[h] for h in headers])

out_wb.save(output_excel)
print(f"\nExcel saved to: {output_excel}")

# === Save to SQLite ===
# Remove existing db if exists
if os.path.exists(output_db):
    os.remove(output_db)

conn = sqlite3.connect(output_db)
cursor = conn.cursor()

# Create table
cursor.execute('''
    CREATE TABLE turnover_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        Debtor TEXT,
        DebtorName TEXT,
        Value REAL,
        Date TEXT,
        DateFixed TEXT,
        Branch TEXT,
        DatePowerBI TEXT
    )
''')

# Insert data
for row_data in all_cleaned_data:
    cursor.execute('''
        INSERT INTO turnover_data (Debtor, DebtorName, Value, Date, DateFixed, Branch, DatePowerBI)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        row_data['Debtor'],
        row_data['Debtor Name'],
        row_data['Value'],
        row_data['Date'].strftime('%Y-%m-%d'),
        row_data['Date Fixed'],
        row_data['Branch'],
        row_data['Date_PowerBI'].strftime('%Y-%m-%d')
    ))

conn.commit()
conn.close()

print(f"SQLite database saved to: {output_db}")
print(f"Total rows: {len(all_cleaned_data)}")
