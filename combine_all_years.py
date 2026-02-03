from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import sqlite3

# Input and output paths
base_folder = "/home/ethan/Desktop/Gork Test/Turn over GABE tuesday update"
output_excel = "/home/ethan/Desktop/Gork Test/all_years_data.xlsx"
output_db = "/home/ethan/Desktop/Gork Test/all_years_data.db"

# Years to process
years = ['2020', '2021', '2022', '2023', '2024', '2025']

def get_branch(filename):
    name = filename.upper().replace('.XLSX', '')
    if 'ALL OVER' in name:
        return None
    return name

def parse_month_code(code):
    """Convert month code like 202401 to datetime"""
    try:
        code_str = str(int(code))
        year = int(code_str[:4])
        month = int(code_str[4:])
        if 1 <= month <= 12:
            return datetime(year, month, 1)
    except:
        pass
    return None

def process_wide_format(sheet, branch):
    """Process wide format files"""
    data = []

    # Find header row to get month columns
    month_columns = {}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        if row and len(row) > 6 and row[3] == 'Debtor':
            # Found header row, get month columns
            for col_idx in range(6, len(row)):
                if row[col_idx]:
                    date_val = parse_month_code(row[col_idx])
                    if date_val:
                        month_columns[col_idx] = date_val
            break

    if not month_columns:
        return data

    # Process data rows
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 6:
            continue

        debtor_code = row[3]
        debtor_name = row[4]

        if not debtor_code or debtor_code == 'Debtor':
            continue
        if not isinstance(debtor_code, str):
            continue

        for col_idx, date_val in month_columns.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    value = float(row[col_idx])
                    if value != 0:
                        data.append({
                            'Debtor': debtor_code,
                            'Debtor Name': debtor_name,
                            'Value': value,
                            'Date': date_val,
                            'Date Fixed': date_val.strftime('%Y-%m-%d'),
                            'Branch': branch,
                            'Date_PowerBI': date_val
                        })
                except (ValueError, TypeError):
                    continue
    return data

all_cleaned_data = []

# Process each year
for year in years:
    year_folder = os.path.join(base_folder, year)
    if not os.path.exists(year_folder):
        print(f"Folder not found: {year_folder}")
        continue

    files = [f for f in os.listdir(year_folder) if f.lower().endswith('.xlsx')]
    files = [f for f in files if 'ALL OVER' not in f.upper()]

    print(f"\n=== Processing {year} ({len(files)} files) ===")

    for filename in files:
        branch = get_branch(filename)
        if branch is None:
            print(f"  Skipping: {filename}")
            continue

        filepath = os.path.join(year_folder, filename)

        try:
            wb = load_workbook(filepath, data_only=True)
            sheet = wb.active

            data = process_wide_format(sheet, branch)
            all_cleaned_data.extend(data)

            print(f"  {filename}: {len(data)} rows")
            wb.close()
        except Exception as e:
            print(f"  Error processing {filename}: {e}")

# Sort by Debtor, then Date
all_cleaned_data.sort(key=lambda x: (x['Debtor'], x['Date']))

# === Save to Excel ===
out_wb = Workbook()
out_sheet = out_wb.active
out_sheet.title = "AllYearsData"

# Same columns as Combined_Analysis_Data.xlsx
headers = ['Debtor', 'Debtor Name', 'Value', 'Date', 'Date Fixed', 'Branch', 'Date_PowerBI']
out_sheet.append(headers)

for row_data in all_cleaned_data:
    out_sheet.append([row_data[h] for h in headers])

out_wb.save(output_excel)
print(f"\nExcel saved to: {output_excel}")

# === Save to SQLite ===
if os.path.exists(output_db):
    os.remove(output_db)

conn = sqlite3.connect(output_db)
cursor = conn.cursor()

cursor.execute('''
    CREATE TABLE turnover_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        Debtor TEXT,
        DebtorName TEXT,
        Value REAL,
        Date TEXT,
        DateFixed TEXT,
        Branch TEXT,
        DatePowerBI TEXT
    )
''')

for row_data in all_cleaned_data:
    cursor.execute('''
        INSERT INTO turnover_data (Debtor, DebtorName, Value, Date, DateFixed, Branch, DatePowerBI)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        row_data['Debtor'],
        row_data['Debtor Name'],
        row_data['Value'],
        row_data['Date'].strftime('%Y-%m-%d'),
        row_data['Date Fixed'],
        row_data['Branch'],
        row_data['Date_PowerBI'].strftime('%Y-%m-%d')
    ))

conn.commit()
conn.close()

print(f"SQLite database saved to: {output_db}")
print(f"\nTotal rows: {len(all_cleaned_data)}")
