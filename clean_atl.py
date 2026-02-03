from openpyxl import load_workbook, Workbook
from datetime import datetime

# Input and output paths
input_file = "/home/ethan/Desktop/Gork Test/2025/ATL.xlsx"
output_file = "/home/ethan/Desktop/Gork Test/2025/ATL_Cleaned.xlsx"

# Load workbook
wb = load_workbook(input_file, data_only=True)
sheet = wb.active

# Month mapping (tuple index to date)
month_map = {
    6: datetime(2025, 1, 1),   # 202501
    7: datetime(2025, 2, 1),   # 202502
    8: datetime(2025, 3, 1),   # 202503
    9: datetime(2025, 4, 1),   # 202504
    10: datetime(2025, 5, 1),  # 202505
    11: datetime(2025, 6, 1),  # 202506
    12: datetime(2025, 7, 1),  # 202507
    13: datetime(2025, 8, 1),  # 202508
    14: datetime(2025, 9, 1),  # 202509
    15: datetime(2025, 10, 1), # 202510
    16: datetime(2025, 11, 1), # 202511
    17: datetime(2025, 12, 1), # 202512
}

cleaned_data = []
branch = "ATL"

# Process rows
for row in sheet.iter_rows(min_row=1, values_only=True):
    if not row or len(row) < 6:
        continue

    debtor_code = row[3]
    debtor_name = row[4]

    if not debtor_code or debtor_code == 'Debtor':
        continue

    if not isinstance(debtor_code, str):
        continue

    # Extract monthly values
    for col_idx, date_val in month_map.items():
        if col_idx < len(row) and row[col_idx] is not None:
            try:
                value = float(row[col_idx])
                if value != 0:
                    cleaned_data.append({
                        'Debtor': debtor_code,
                        'Debtor Name': debtor_name,
                        'Value': value,
                        'Date': date_val,
                        'Date Fixed': date_val.strftime('%Y-%m-%d'),
                        'Branch': branch,
                        'Date_PowerBI': date_val
                    })
            except (ValueError, TypeError):
                continue

# Sort by Debtor, then by Date
cleaned_data.sort(key=lambda x: (x['Debtor'], x['Date']))

# Create new Excel workbook
out_wb = Workbook()
out_sheet = out_wb.active
out_sheet.title = "ATL_Cleaned"

# Write header
headers = ['Debtor', 'Debtor Name', 'Value', 'Date', 'Date Fixed', 'Branch', 'Date_PowerBI']
out_sheet.append(headers)

# Write data
for row_data in cleaned_data:
    out_sheet.append([row_data[h] for h in headers])

# Save
out_wb.save(output_file)

print(f"Cleaned data saved to: {output_file}")
print(f"Total rows: {len(cleaned_data)}")
